# -*- coding: utf-8 -*-
"""선명(entity 2·3) 2026-07 수금 적재 SQL 생성기 (2026-08-07).

★★ **은행연동 중복 배제가 이 스크립트의 존재 이유다.**
   동산 7월 이관에서 같은 사고가 났다 — 이관 **전에** MES 에 이미 있던 3건(1,227,380)을
   확인해 놓고도 배제를 안 넣어 이중 계상됐다. **총액 검산으로는 안 잡힌다**(이카운트 합계는 맞았다).
   자동매칭(`[은행연동]`)은 계속 돌고 있어 **이관 시점마다 겹침이 새로 생긴다.**
   ⇒ (거래처 · 일자 · 금액) 이 겹치는 기존 `payments` 가 있으면 **그 건은 넣지 않는다.**

★ 수금 파일엔 **사업자번호가 없다**(일자-No. · 거래처명 · 금액 · 적요 4열뿐).
   그래서 판매 파일의 **거래처명 → 사업자번호** 를 경유해 2단계로 붙인다(동산 7월과 같은 방식).

★ entity = 판매와 같은 규칙(담당 최인호 → 3). 수금 파일엔 담당이 없으므로
   **거래처가 7월 판매에서 어느 법인이었는지**로 정한다. 판매가 없는 거래처는 2(선명).

사용: python gen_e2_receipts_sql.py <clients.json> <sales.xlsx> <receipts.xlsx> <existing_payments.json>
산출: load/payments_2607_e2.sql · load/rollback_payments_2607_e2.sql
"""
import openpyxl, json, sys, re, os, warnings
from collections import Counter
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

CLIENTS_JSON, SALES_XLSX, RCP_XLSX, PAY_JSON = sys.argv[1:5]
OUT = os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'dongsan-import', 'load')
MARKER = '선명 이관 2026-07 수금'
LINE = re.compile(r'^(\d{4})/(\d\d)/(\d\d)\s+-(\d+)$')
PAREN = re.compile(r'\([^)]*\)')

sys.path.insert(0, os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'dongsan-import'))
from normalize_client_code import normalize  # noqa


def nname(s):
    s = PAREN.sub('', str(s or ''))
    for t in ('주식회사', '(주)', '㈜', '유한회사'):
        s = s.replace(t, '')
    return re.sub(r'\s+', '', s).strip()


def q(s):
    return str(s or '').replace("'", "''")


cl = json.load(open(CLIENTS_JSON, encoding='utf-8'))
BY_CODE = {str(c['client_code']).strip(): c['id'] for c in cl if c['client_code']}
BY_BRN = {str(c['brn']).strip(): c['id'] for c in cl if c['brn']}
BY_NAME = {}
for c in cl:
    BY_NAME.setdefault(nname(c['client_name']), c['id'])

# ── 판매 파일에서 거래처명 → (client_id, entity) 경유표 ──────────────────────
wb = openpyxl.load_workbook(SALES_XLSX)
ws = wb.active
name2brn, name2ent = {}, {}
for r in ws.iter_rows(min_row=3, values_only=True):
    if not LINE.match(str(r[0] or '').strip()):
        continue
    nm = nname(r[2])
    name2brn.setdefault(nm, str(r[1] or '').strip())
    name2ent.setdefault(nm, 3 if str(r[3] or '').strip() == '최인호' else 2)
wb.close()


# 표기 차이 별칭 — 이카운트 ↔ MES. 「애/에」 한 글자 차이라 이름 정규화가 못 잡는다.
NAME_ALIAS = {'연안애드마트': '연안에드마트'}


def resolve(name):
    nm = nname(name)
    nm = NAME_ALIAS.get(nm, nm)
    brn = name2brn.get(nm)
    for cand in (brn, (normalize(brn)[0] if brn else None)):
        if cand and cand in BY_CODE:
            return BY_CODE[cand], name2ent.get(nm, 2)
        if cand and cand in BY_BRN:
            return BY_BRN[cand], name2ent.get(nm, 2)
    if nm in BY_NAME:
        return BY_NAME[nm], name2ent.get(nm, 2)
    return None, 2


# ★ **entity 를 반드시 키에 넣는다.** (거래처·일자·금액) 만 보면 **동산(e1) 수금을 선명 것으로 오인**한다 —
#   거래처는 법인 공유라 같은 곳이 두 법인에 동시에 입금할 수 있고, 실제로 그런 건이 있었다
#   (2026-07-09 인효교재악기 9,837,520 = 동산 이관분인데 선명 수금을 중복으로 빼버렸다).
#   중복 대상은 **같은 법인(2·3)** 안의 은행연동분뿐이다.
existing = {(p['entity_id'], p['client_id'], str(p['payment_date'])[:10], round(float(p['amount'])))
            for p in json.load(open(PAY_JSON, encoding='utf-8'))}

wb = openpyxl.load_workbook(RCP_XLSX)
ws = wb.active
rows, unresolved = [], []
for r in ws.iter_rows(min_row=3, values_only=True):
    m = LINE.match(str(r[0] or '').strip())
    if not m:
        continue
    date = '%s-%s-%s' % (m.group(1), m.group(2), m.group(3))
    cid, ent = resolve(r[1])
    if cid is None:
        unresolved.append((date, str(r[1]), r[2]))
        continue
    rows.append({'date': date, 'no': int(m.group(4)), 'cid': cid, 'ent': ent,
                 'amt': float(r[2] or 0), 'memo': str(r[3] or ''), 'name': str(r[1] or '')})
wb.close()

dup = [x for x in rows if (x['ent'], x['cid'], x['date'], round(x['amt'])) in existing]
keep = [x for x in rows if (x['ent'], x['cid'], x['date'], round(x['amt'])) not in existing]

buf, ent_cnt = [], Counter()
for x in keep:
    ts = '%s 09:00:00' % x['date']
    notes = '%s %s-%d%s' % (MARKER, x['date'], x['no'], (' · ' + x['memo']) if x['memo'] else '')
    ent_cnt[x['ent']] += 1
    buf.append(
        "INSERT INTO payments (client_id,entity_id,payment_date,amount,payment_method,notes,created_by,created_at) "
        "VALUES (%d,%d,'%s',%s,'BANK_TRANSFER','%s',5,'%s');"
        % (x['cid'], x['ent'], x['date'], repr(x['amt']) if x['amt'] % 1 else str(int(x['amt'])),
           q(notes), ts))

open(os.path.join(OUT, 'payments_2607_e2.sql'), 'w', encoding='utf-8').write(
    '-- 선명(entity 2·3) 2026-07 수금 적재 (2026-08-07)\n'
    '-- 생성기 = docs/dongsan-import/gen_e2_receipts_sql.py\n'
    '-- ★ 은행연동 중복 %d건 배제 완료 (거래처·일자·금액 일치 기준)\n'
    '-- 롤백 = load/rollback_payments_2607_e2.sql\n\n' % len(dup) + '\n'.join(buf) + '\n')
open(os.path.join(OUT, 'rollback_payments_2607_e2.sql'), 'w', encoding='utf-8').write(
    "-- 선명 2026-07 수금 적재 롤백 (2026-08-07)\n"
    "DELETE FROM payments WHERE notes LIKE '%s%%';\n" % MARKER)

print('수금 파일 %d건 · 합계 %s' % (len(rows) + len(unresolved), format(int(sum(x['amt'] for x in rows)), ',')))
print('  ★ 은행연동 중복 배제 %d건 %s' % (len(dup), format(int(sum(x['amt'] for x in dup)), ',')))
for x in dup:
    print('      %s | %s | %s' % (x['date'], x['name'], format(int(x['amt']), ',')))
print('  적재 대상 %d건 (e2 %d · e3 %d) 합계 %s'
      % (len(keep), ent_cnt[2], ent_cnt[3], format(int(sum(x['amt'] for x in keep)), ',')))
print('  거래처 미해결 %d건' % len(unresolved))
for d, n, a in unresolved:
    print('      %s | %s | %s' % (d, n, a))
