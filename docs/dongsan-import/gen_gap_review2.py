# -*- coding: utf-8 -*-
"""갭 워크시트 2차 — **번호로 답하는** 검토표 (2026-08-09).

★ 1차에서 배운 것
  `★확정코드` 열에 코드를 적으시라고 냈는데, 실제로는 **번호(1/2/3)·「신규」·「x」**로 채워 주셨다.
  그게 훨씬 빠르고 오타도 안 난다 — **열 설계가 틀렸던 것**이라 이번 판은 그 형식으로 낸다.

  `★선택` 에 다음 중 하나만 적으면 된다:
    1 · 2 · 3  후보 채택      N  신설 필요      X  사용 안 함(연결 불필요)
  후보1 점수가 높은 행은 **1 을 미리 채워 뒀다** — 맞으면 그대로 두고 다르면 고치면 된다.

★ 계열로 묶었다
  1차에선 「채널바 옆마구리」 20행이 흩어져 나와 같은 판단을 스무 번 하게 만들었다.
  이번엔 **품명 기준으로 묶고 그룹 번호를 매긴다.** 같은 그룹은 대개 답이 같으므로
  대표 행 하나만 보고 나머지는 복사하면 된다(규격만 다른 형제들이다).

★ 이미 확정된 것은 뺀다
  1차에서 답 주신 24건과 「기존연결 유력」은 대상이 아니다 — **판단이 필요한 것만** 남긴다.

사용: python docs/dongsan-import/gen_gap_review2.py <gap_v2.xlsx>
산출: docs/analysis/2026-08-09-gap-review2.xlsx
"""
import openpyxl, sys, os, re, json, warnings
from collections import defaultdict
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

SRC = sys.argv[1] if len(sys.argv) > 1 else 'docs/analysis/2026-08-08-ecount-item-gap-v2.xlsx'
OUT = os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'analysis', '2026-08-09-gap-review2.xlsx')

ws = openpyxl.load_workbook(SRC).active
H = [c.value for c in ws[1]]
ix = {n: H.index(n) for n in H if n}

# ★ **품목맵에 이미 있으면 뺀다.** 1차 답변이 계열 전체로 번지는 경우가 있어서다 —
#   「후판 = 간판 후판」이라고 한 규격에 답하시면 50~180cm **전 규격이 확정**된 것이고,
#   그걸 워크시트가 모르면 같은 질문을 다섯 번 더 한다(실제로 그랬다).
#   품목맵은 그 확정을 이미 담고 있으므로 그걸 기준으로 거른다.
MAP = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  'e2_purchase_item_map.json'), encoding='utf-8'))
def resolved(nm, sp):
    nm, sp = str(nm or '').strip(), str(sp or '').strip()
    return ('%s [%s]|' % (nm, sp)) in MAP or ('%s|%s' % (nm, sp)) in MAP or ('%s|' % nm) in MAP

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[ix['자동판단']] != '검토':
        continue
    if str(r[ix['★확정코드(기입)']] or '').strip():   # 1차에서 이미 답 주신 것
        continue
    if resolved(r[ix['품목명']], r[ix['규격']]):
        continue
    rows.append(r)

# ── 품명 계열로 묶는다. 계열 = 품명에서 숫자·규격 흔적을 걷어낸 것
NUM = re.compile(r'[\d.]+\s*(mm|cm|폭|T|W|w|파이|호)?')
def family(nm):
    return NUM.sub('', str(nm)).strip(' -()[]') or str(nm)

g = defaultdict(list)
for r in rows:
    g[family(r[ix['품목명']])].append(r)

# 큰 계열 먼저 — 한 번 판단해 여러 행을 처리할 수 있는 것부터 보시게
order = sorted(g.items(), key=lambda kv: (-len(kv[1]), kv[0]))

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
wb = openpyxl.Workbook()
sh = wb.active
sh.title = '검토2'

sh.append(['★선택 기입법 →  1/2/3 = 후보 채택   ·   N = 신설 필요   ·   X = 사용 안 함',
           '', '', '', '', '', '', '', '', '', ''])
sh.append(['같은 그룹은 규격만 다른 형제입니다 — 대표 행 하나만 보고 나머지는 같은 값을 넣으시면 됩니다.',
           '', '', '', '', '', '', '', '', '', ''])
sh.append([])
HDR = ['그룹', '★선택', '이카운트 품목명', '규격', '구분', '실적',
       '후보1', '후보1 품목명', '점수', '후보2', '후보3']
sh.append(HDR)
HROW = sh.max_row

n = 0
for gi, (fam, items) in enumerate(order, 1):
    for r in sorted(items, key=lambda x: str(x[ix['규격']])):
        sc = float(r[ix['점수1']] or 0)
        sh.append([
            gi,
            1 if sc >= 0.6 else '',                 # 높으면 미리 채운다
            r[ix['품목명']], r[ix['규격']], r[ix['구분']], r[ix['우선순위']],
            r[ix['후보1코드']], r[ix['후보1명']], sc,
            r[ix['후보2코드']] or '', r[ix['후보3코드']] or '',
        ])
        n += 1

# ── 서식
for c in sh[HROW]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
sh['A1'].font = Font(bold=True, size=12)
sh['A2'].font = Font(italic=True, color='555555')

thin = Side(style='thin', color='BFBFBF')
prev = None
for i in range(HROW + 1, sh.max_row + 1):
    gi = sh.cell(i, 1).value
    # 그룹이 바뀌는 자리에 굵은 선 — 눈으로 계열 경계가 보이게
    if prev is not None and gi != prev:
        for j in range(1, len(HDR) + 1):
            sh.cell(i, j).border = Border(top=Side(style='medium', color='4472C4'))
    prev = gi
    if gi % 2 == 0:
        for j in range(1, len(HDR) + 1):
            sh.cell(i, j).fill = PatternFill('solid', fgColor='F2F2F2')
    sh.cell(i, 2).fill = PatternFill('solid', fgColor='FFF2CC')   # ★선택 열 강조
    sh.cell(i, 2).font = Font(bold=True)
    sh.cell(i, 2).alignment = Alignment(horizontal='center')
    for j in range(1, len(HDR) + 1):
        cur = sh.cell(i, j).border
        sh.cell(i, j).border = Border(top=cur.top, bottom=thin)

for col, w in zip('ABCDEFGHIJK', [6, 8, 32, 20, 8, 12, 18, 30, 7, 18, 18]):
    sh.column_dimensions[col].width = w
sh.freeze_panes = 'C%d' % (HROW + 1)
sh.auto_filter.ref = 'A%d:K%d' % (HROW, sh.max_row)
wb.save(OUT)

print('검토 대상 %d행 · %d계열 → %s' % (n, len(order), OUT))
print('미리 1 채운 행: %d (점수 0.6+)' % sum(1 for _, its in order for r in its if float(r[ix['점수1']] or 0) >= 0.6))
print('\n계열 상위:')
for fam, its in order[:10]:
    print('   %2d행  %s' % (len(its), fam[:44]))
