# -*- coding: utf-8 -*-
"""동산기획 이카운트 수금현황 2601~2606 → MES payments 적재 SQL 생성기.

⚠️ 수금현황에는 거래처코드가 없다(일자-No.·거래처명·금액·적요) → 이름 매칭.
   정확성 담보 = **거래처별 수금 합계를 채권파일(거래처코드 보유)과 전수 대사** — 이름 오매핑이 있으면 여기서 터진다.
전례 = docs/sunmyung-import/08_payments.sql (payments INSERT·계좌이체·마커 notes).
제외 = 선명커뮤니케이션(342-86-03531, 매출 제외와 대칭 — INSPECTION §④ 결정).

사용: python gen_payments_sql.py <clients_ids.json>
"""
import openpyxl, json, sys, re, os, hashlib, csv
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
ROOT = r'C:\Users\user\dongsan_mes'
SRC = os.path.join(ROOT, '품목마스터', '원천주문데이터')
OUT = os.path.join(ROOT, 'docs', 'dongsan-import', 'load')
os.makedirs(OUT, exist_ok=True)
sys.path.insert(0, os.path.join(ROOT, 'docs', 'dongsan-import'))
from normalize_client_code import normalize  # noqa

CLIENTS_JSON = sys.argv[1]
LINE = re.compile(r'^\d{4}/\d\d/\d\d\s+-\d+$')
EXCL_CODES = {'342-86-03531', '314-81-84311'}
RRN = re.compile(r'^\d{6}-\d{7}$')
MARKER = '동산 이카운트 이관 2026-07-30 수금'


def nrm(s):
    """이름 정규화 — '(사용중단)' 리포트 접미 제거 + 공백 축약."""
    s = re.sub(r'\(사용중단\)\s*$', '', str(s or '').strip())
    return re.sub(r'\s+', ' ', s)


def key(s):
    return re.sub(r'\s+', '', nrm(s))


# ── ① 이카운트 마스터: 이름 → 코드 (충돌 추적) ─────────────────────────
wb = openpyxl.load_workbook(os.path.join(SRC, '동산_거래처등록_20260730_16col.xlsx'), data_only=True)
ws = wb.worksheets[0]
name2codes = defaultdict(set)
for r in ws.iter_rows(min_row=3, max_col=2, values_only=True):
    code, name = str(r[0] or '').strip(), nrm(r[1])
    if code and name and not code.startswith('회사명'):
        name2codes[key(name)].add(code)
wb.close()
print(f'마스터 이름 {len(name2codes):,}개 · 동명 충돌 {sum(1 for v in name2codes.values() if len(v) > 1)}개')

# ── ② 채권파일: 코드 → 수금합계 (검증 정본) + 이름 보조 ─────────────────
wb = openpyxl.load_workbook(os.path.join(SRC, '동산_거래처별채권_2601-2606.xlsx'), data_only=True)
ws = wb.worksheets[0]
ar = {}
ar_name2code = defaultdict(set)
for r in ws.iter_rows(min_row=3, max_col=8, values_only=True):
    code, name = str(r[0] or '').strip(), nrm(r[1])
    if not code or not name or ' 계' in code or code.startswith('총') or '합계' in code:
        continue
    vals = [v if isinstance(v, (int, float)) else 0 for v in r[2:8]]
    ar[code] = dict(name=name, opening=vals[0], sales=vals[1], acct=vals[2], collected=vals[3],
                    diff=vals[4], balance=vals[5])
    ar_name2code[key(name)].add(code)
wb.close()
tot_col = sum(v['collected'] for v in ar.values())
print(f'채권파일 거래처 {len(ar):,} · 수금합계 {tot_col:,.0f} (기대 3,043,174,438)')
assert int(round(tot_col)) == 3043174438, '채권파일 수금합계 불일치'

# ── ③ MES client 해석 ─────────────────────────────────────────────────
_cl = json.load(open(CLIENTS_JSON, encoding='utf-8-sig'))[0]['results']
BY_CODE = {str(c['client_code']).strip(): c['id'] for c in _cl if c['client_code']}
BY_BRN = {str(c['brn']).strip(): c['id'] for c in _cl if c['brn']}
_reg = list(csv.DictReader(open(os.path.join(ROOT, 'docs', 'dongsan-import', '점검1b_미등록거래처_등록용.csv'),
                                encoding='utf-8-sig')))
BY_SRC = {r['이카운트코드']: r['client_code'] for r in _reg if not r['이카운트코드'].startswith('(')}
BY_HASH = {r['원본코드해시']: r['client_code'] for r in _reg if r['원본코드해시']}


def resolve(cc):
    if cc in BY_CODE:
        return BY_CODE[cc]
    if cc in BY_BRN:
        return BY_BRN[cc]
    ncc, _, _, _ = normalize(cc)
    if ncc and ncc in BY_CODE:
        return BY_CODE[ncc]
    if ncc and ncc in BY_BRN:
        return BY_BRN[ncc]
    if cc in BY_SRC:
        return BY_CODE.get(BY_SRC[cc])
    if RRN.match(cc):
        h = 'sha256:' + hashlib.sha256(re.sub(r'\D', '', cc).encode()).hexdigest()[:16]
        if h in BY_HASH:
            return BY_CODE.get(BY_HASH[h])
    return None


def name_to_code(name):
    """이름 → 이카운트 코드. 동명 충돌은 채권파일 등장 코드 우선."""
    k = key(name)
    cands = name2codes.get(k, set()) | ar_name2code.get(k, set())
    if not cands:
        return None, 'no-name'
    if len(cands) == 1:
        return next(iter(cands)), 'ok'
    in_ar = [c for c in cands if c in ar and ar[c]['collected']]
    if len(in_ar) == 1:
        return in_ar[0], 'ar-disambig'
    return None, f'ambiguous:{sorted(cands)}'


# ── ④ 수금현황 로드 ───────────────────────────────────────────────────
wb = openpyxl.load_workbook(os.path.join(SRC, '동산_수금현황_2601-2606.xlsx'), data_only=True)
ws = wb.worksheets[0]
rows = []
for r in ws.iter_rows(min_row=3, max_col=4, values_only=True):
    head = str(r[0] or '').strip()
    if not LINE.match(head):
        continue
    d8, no = head.split()
    amt = r[2] if isinstance(r[2], (int, float)) else 0
    rows.append(dict(date=d8.replace('/', '-'), no=no, name=nrm(r[1]), amt=amt,
                     memo=str(r[3] or '').strip()))
wb.close()
tot = sum(x['amt'] for x in rows)
print(f'수금 라인 {len(rows):,} · 합계 {tot:,.0f}')
assert int(round(tot)) == 3043174438, f'수금현황 합계 불일치: {tot}'

# ── ⑤ 이름 → 코드 → client_id + 제외 ──────────────────────────────────
fails = defaultdict(lambda: [0, 0])
excl = [0, 0]
target = []
amb = defaultdict(list)


def accept(x, code):
    if code in EXCL_CODES:
        excl[0] += 1
        excl[1] += x['amt']
        return
    cid = resolve(code)
    if cid is None:
        fails[(x['name'], f'no-client:{code}')][0] += 1
        fails[(x['name'], f'no-client:{code}')][1] += x['amt']
        return
    x['code'], x['cid'] = code, cid
    target.append(x)


for x in rows:
    code, how = name_to_code(x['name'])
    if code is None and how.startswith('ambiguous'):
        amb[key(x['name'])].append(x)
        continue
    if code is None:
        fails[(x['name'], how)][0] += 1
        fails[(x['name'], how)][1] += x['amt']
        continue
    accept(x, code)

# 동명 다코드 → 채권파일 목표 금액으로 라인 분배 (각 단계 정확·유일 일치 + 잔여 0 필수)
for k, lines in amb.items():
    cands = sorted(name2codes.get(k, set()) | ar_name2code.get(k, set()))
    remain = {c: ar.get(c, {}).get('collected', 0) for c in cands}
    plan, ok = [], True
    for x in sorted(lines, key=lambda v: -abs(v['amt'])):
        exact = [c for c in cands if abs(remain[c] - x['amt']) < 0.01]
        if len(exact) != 1:
            ok = False
            break
        remain[exact[0]] -= x['amt']
        plan.append((x, exact[0]))
    if ok and all(abs(v) < 0.01 for v in remain.values()):
        for x, c in plan:
            accept(x, c)
        print(f'동명 분배 OK: {lines[0]["name"]} → ' + ' · '.join(f'{c}={ar[c]["collected"]:,.0f}' for c in cands))
    else:
        for x in lines:
            fails[(x['name'], f'ambiguous:{cands}')][0] += 1
            fails[(x['name'], f'ambiguous:{cands}')][1] += x['amt']
print(f'제외(선명·자기거래): {excl[0]}건 {excl[1]:,.0f}')
print(f'적재 대상: {len(target):,}건 {sum(x["amt"] for x in target):,.0f}')
if fails:
    print(f'★미해석 {len(fails)}건:')
    for (nm, how), v in sorted(fails.items(), key=lambda x: -x[1][1]):
        print(f'  {nm[:30]:<32}{how[:44]:<46}{v[0]:>4}건 {v[1]:>13,.0f}')

# ── ⑥ ★거래처별 전수 대사 — 채권파일 수금합계와 대조 ───────────────────
by_code = defaultdict(float)
for x in target:
    by_code[x['code']] += x['amt']
mis = []
for code, s in by_code.items():
    want = ar.get(code, {}).get('collected', 0)
    if abs(s - want) > 0.01:
        mis.append((code, ar.get(code, {}).get('name', '?'), s, want))
# 역방향: 채권파일에 수금이 있는데 수금현황 매핑에 없는 코드 (제외 코드 제하고)
rev = [(c, v['name'], v['collected']) for c, v in ar.items()
       if v['collected'] and c not in by_code and c not in EXCL_CODES]
print(f'\n★채권파일 대사: 불일치 {len(mis)}건 · 역방향 누락 {len(rev)}건')
for c, n, s, w in sorted(mis, key=lambda x: -abs(x[2] - x[3]))[:10]:
    print(f'  {c:<16}{n[:24]:<26}수금현황 {s:>13,.0f} vs 채권 {w:>13,.0f} (차 {s-w:+,.0f})')
for c, n, w in sorted(rev, key=lambda x: -x[2])[:10]:
    print(f'  [역방향] {c:<16}{n[:24]:<26}채권 수금 {w:>13,.0f} — 수금현황 매핑 없음')

if '--audit' in sys.argv:
    sys.exit(0 if not mis and not rev and not fails else 1)

assert not mis and not rev and not fails, '대사 실패 — SQL 미생성'

# ── ⑦ SQL 생성 ────────────────────────────────────────────────────────
def q(s):
    return re.sub(r'[\x00-\x08\x0b-\x1f]', ' ', str(s or '').replace("'", "''"))


def sqlnum(v):
    return repr(round(v, 4)) if isinstance(v, float) and not float(v).is_integer() else str(int(v))


def method(memo):
    if '어음' in memo:
        return '어음'
    if '현금' in memo:
        return '현금'
    if '카드' in memo:
        return '카드'
    return '계좌이체'


buf = []
for x in sorted(target, key=lambda x: (x['date'], int(x['no']))):
    ts = f"{x['date']} 09:00:00"
    notes = MARKER + (f" · {x['memo']}" if x['memo'] else '')
    buf.append(
        "INSERT INTO payments (client_id,payment_date,amount,payment_method,reference_number,"
        "notes,created_by,entity_id,created_at,updated_at) VALUES "
        f"({x['cid']},'{x['date']}',{sqlnum(x['amt'])},'{method(x['memo'])}',"
        f"'{x['date']} {x['no']}','{q(notes)}',5,1,'{ts}','{ts}');")

# 0.8MB 조각 (파일축 정책)
parts, part, size, idx = [], [], 0, 0
for l in buf:
    b = len(l.encode('utf-8')) + 1
    if size + b > 800_000:
        idx += 1
        parts.append((f'payments_p{idx:02d}.sql', part))
        part, size = [], 0
    part.append(l)
    size += b
idx += 1
parts.append((f'payments_p{idx:02d}.sql', part))
for name, lines in parts:
    with open(os.path.join(OUT, name), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')
    print(f'{name}: {len(lines):,}문장')
with open(os.path.join(OUT, 'rollback_payments.sql'), 'w', encoding='utf-8') as f:
    f.write(f"DELETE FROM payments WHERE entity_id=1 AND notes LIKE '{MARKER}%';\n")

print(f'\n적재 SQL {len(buf):,}문장 · 합계 {sum(x["amt"] for x in target):,.0f} · 조각 {len(parts)}개 → {OUT}')
