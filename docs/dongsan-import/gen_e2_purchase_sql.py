# -*- coding: utf-8 -*-
"""선명(entity 2·3) 2026-07 **매입** 적재 SQL 생성기 (2026-08-08).

판매 생성기(`gen_e2_sales_sql.py`)의 짝이다. 공유하는 건 전표 묶기·거래처 해석·마커 규칙이고,
적재 대상이 다르다 — `purchase_orders` + `purchase_order_items` (청구그룹 없음).

★ 왜 7월 매입만 따로 도는가
  선명 매입은 **2026-06-30 에서 끊겨** 있었다. 7월 판매·수금은 이관됐는데 매입만 빠져서
  「EP1852 를 7월에 팔았는데 매입이 없다」 같은 **품목 단위 오진**을 유발했다(실제론 월 전체 누락).

★ 번호 = `SMP-2607-###` (날짜 기반)
  기존 이관분이 `SMP-0001`~`SMP-0248` 연번이라, 연번을 이어 붙이면 **재실행·부분복구 때 충돌**한다.
  월을 번호에 박으면 언제 들어온 건지 번호만 봐도 알 수 있고 겹칠 수가 없다.

★ 마커 = 월·법인 분리. 롤백이 `notes LIKE '{MARKER}%'` 라 같은 마커를 쓰면
  **기존 적재분까지 지운다**(동산에서 상반기 7,642건이 날아갈 뻔했다).

★ entity 배정 = 판매와 같은 규칙(담당자). 다만 매입은 실제로 거의 전량 선명이다 —
  6월까지 SMP-* 248건이 **전부 entity 2** 다. 담당 컬럼이 없으면 2로 간다.

★ 품목 미매칭은 `item_id NULL`. 금액·거래처·전표는 정확하고 원가 분석에서만 빠진다.
  판매 쪽에서 확인됐듯 **품목은 나중에 붙일 수 있지만 전표는 그때 다시 못 받는다** —
  매칭을 기다리느라 적재를 미루지 않는다.

사용: python docs/dongsan-import/gen_e2_purchase_sql.py <clients.json> <items.json> <map.json> <purchase.xlsx>
산출: load/purchase_2607_e2.sql · load/rollback_purchase_2607_e2.sql
"""
import openpyxl, json, sys, re, os, warnings
from collections import defaultdict, Counter
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

CLIENTS_JSON, ITEMS_JSON, MAP_JSON, XLSX = sys.argv[1:5]
OUT = os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'dongsan-import', 'load')
MARKER = '선명 이관 매입 2026-07'
CJ_REP = '최인호'
LINE = re.compile(r'^(\d{4})/(\d\d)/(\d\d)\s+-(\d+)$')

sys.path.insert(0, os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'dongsan-import'))
from normalize_client_code import normalize  # noqa

cl = json.load(open(CLIENTS_JSON, encoding='utf-8'))
BY_CODE = {str(c['client_code']).strip(): c['id'] for c in cl if c['client_code']}
BY_BRN = {str(c['brn']).strip(): c['id'] for c in cl if c['brn']}
items = {i['id']: i for i in json.load(open(ITEMS_JSON, encoding='utf-8'))}
IMAP = json.load(open(MAP_JSON, encoding='utf-8'))


def q(s):
    return str(s or '').replace("'", "''")


def num(v):
    if v is None or str(v).strip() == '':
        return 0
    try:
        return float(v)
    except ValueError:
        return 0


def sqlnum(v):
    return str(int(v)) if float(v) == int(v) else repr(float(v))


def resolve_client(code):
    c = str(code or '').strip()
    for d in (BY_CODE, BY_BRN):
        if c in d:
            return d[c]
    norm, _, _, _ = normalize(c)
    for d in (BY_CODE, BY_BRN):
        if norm and norm in d:
            return d[norm]
    return None


# ── 헤더에서 컬럼 위치를 **읽는다**. 판매현황과 열 순서가 같다는 보장이 없다.
#    위치를 상수로 박으면 열이 하나 밀렸을 때 조용히 다른 값이 들어간다.
COLS = {'일자': None, '거래처코드': None, '거래처명': None, '사원': None, '품목명': None,
        '규격': None, '내용': None, '수량': None, '단가': None, '공급가': None, '부가세': None}
ALIAS = {'일자-No.': '일자', '거래처코드': '거래처코드', '거래처명': '거래처명',
         '사원(담당)명': '사원', '담당자명': '사원', '품목명': '품목명', '규격': '규격',
         '내용': '내용', '수량': '수량', '단가': '단가', '공급가액': '공급가', '부가세': '부가세'}

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
hdr_row = None
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    hits = {}
    for ci, cell in enumerate(row):
        key = ALIAS.get(str(cell or '').strip())
        if key:
            hits[key] = ci
    if '일자' in hits and '품목명' in hits:
        COLS.update(hits)
        hdr_row = ri
        break
assert hdr_row, '헤더 행을 못 찾았다 — 조회 구분이 「라인별」인지 확인할 것(일별이면 일합계만 나온다)'
missing = [k for k, v in COLS.items() if v is None]
assert not missing, '헤더에 없는 컬럼: %s' % missing


def cell(r, key):
    return r[COLS[key]]


vouchers = defaultdict(lambda: {'lines': [], 'rep': '', 'cli': None, 'name': ''})
for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
    m = LINE.match(str(cell(r, '일자') or '').strip())
    if not m:
        continue
    key = ('%s-%s-%s' % (m.group(1), m.group(2), m.group(3)), int(m.group(4)))
    v = vouchers[key]
    v['rep'] = v['rep'] or str(cell(r, '사원') or '').strip()
    v['cli'] = v['cli'] or resolve_client(cell(r, '거래처코드'))
    v['name'] = v['name'] or str(cell(r, '거래처명') or '')
    v['lines'].append({'nm': str(cell(r, '품목명') or ''), 'sp': str(cell(r, '규격') or ''),
                       'ct': str(cell(r, '내용') or ''), 'qty': num(cell(r, '수량')),
                       'up': num(cell(r, '단가')), 'sup': num(cell(r, '공급가')),
                       'vat': num(cell(r, '부가세'))})
wb.close()

bad = [(k, v['name']) for k, v in vouchers.items() if not v['cli']]
assert not bad, '공급처 미해결 전표: %s' % bad[:5]

buf, ent_cnt, sup_tot, vat_tot, lines_n, miss_n, miss_amt = [], Counter(), 0, 0, 0, 0, 0
for seq, ((date, no), v) in enumerate(sorted(vouchers.items()), 1):
    ent = 3 if v['rep'] == CJ_REP else 2
    tot = sum(l['sup'] for l in v['lines'])
    vat = sum(l['vat'] for l in v['lines'])
    fin = tot + vat
    pnum = 'SMP-2607-%03d' % seq
    ts = '%s 09:00:00' % date
    notes = '%s %s-%d%s' % (MARKER, date, no, (' · 담당 ' + v['rep']) if v['rep'] else '')
    ent_cnt[ent] += 1
    sup_tot += tot
    vat_tot += vat
    buf.append(
        "INSERT INTO purchase_orders (po_number,supplier_id,status,order_date,total_amount,"
        "vat_amount,final_amount,notes,created_by,entity_id,created_at,updated_at) VALUES "
        "('%s',%d,'RECEIVED','%s',%s,%s,%s,'%s',5,%d,'%s','%s');"
        % (pnum, v['cli'], date, sqlnum(tot), sqlnum(vat), sqlnum(fin), q(notes), ent, ts, ts))
    for so, l in enumerate(v['lines']):
        iid = IMAP.get('%s|%s' % (l['nm'].strip(), l['sp'].strip()))
        if iid is None:
            miss_n += 1
            miss_amt += l['sup']
        unit = (items.get(iid, {}).get('unit') or 'EA') if iid else 'EA'
        # received_quantity = quantity : 이관분은 **이미 입고된 과거 거래**다. 0 으로 두면
        #   미입고 발주로 잡혀 입고대기 화면이 과거 전표로 가득 찬다.
        buf.append(
            "INSERT INTO purchase_order_items (po_id,item_id,item_name,category_name,quantity,"
            "received_quantity,accepted_quantity,unit,unit_price,amount,vat_included,sort_order,"
            "line_status,notes,created_at,updated_at) VALUES "
            "((SELECT id FROM purchase_orders WHERE po_number='%s'),%s,'%s','%s',%s,%s,%s,'%s',%s,%s,0,%d,"
            "'RECEIVED','%s','%s','%s');"
            % (pnum, iid if iid else 'NULL', q(l['nm']), q(l['sp']),
               sqlnum(l['qty']), sqlnum(l['qty']), sqlnum(l['qty']), q(unit),
               sqlnum(l['up']), sqlnum(l['sup']), so, q(l['ct']), ts, ts))
        lines_n += 1

open(os.path.join(OUT, 'purchase_2607_e2.sql'), 'w', encoding='utf-8').write(
    '-- 선명(entity 2·3) 2026-07 매입 적재 (2026-08-08)\n'
    '-- 생성기 = docs/dongsan-import/gen_e2_purchase_sql.py (근거는 그 파일 헤더)\n'
    '-- 번호 = SMP-2607-### (날짜 기반) — 기존 SMP-0001~0248 연번과 겹치지 않는다\n'
    '-- 롤백 = load/rollback_purchase_2607_e2.sql\n\n' + '\n'.join(buf) + '\n')
open(os.path.join(OUT, 'rollback_purchase_2607_e2.sql'), 'w', encoding='utf-8').write(
    "-- 선명 2026-07 매입 적재 롤백 (2026-08-08)\n"
    "-- ⚠️ 마커가 '%s' 로 **월·법인 분리**돼 있다. 다른 이관분은 안 지워진다.\n"
    "DELETE FROM purchase_order_items WHERE po_id IN (SELECT id FROM purchase_orders WHERE notes LIKE '%s%%');\n"
    "DELETE FROM purchase_orders WHERE notes LIKE '%s%%';\n" % (MARKER, MARKER, MARKER))

print('전표 %d (e2 %d · e3 %d) · 라인 %d · 공급가 %s · 부가세 %s · 합계 %s'
      % (len(vouchers), ent_cnt[2], ent_cnt[3], lines_n,
         format(int(sup_tot), ','), format(int(vat_tot), ','), format(int(sup_tot + vat_tot), ',')))
print('품목 미매칭 %d라인 %s원 (item_id NULL — 나중에 붙인다)' % (miss_n, format(int(miss_amt), ',')))
