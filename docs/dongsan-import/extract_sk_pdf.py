# -*- coding: utf-8 -*-
"""서울경금속 명세 PDF 6개 → 상세 라인 추출 + 3중 검산 + 엑셀.
검산: ①행 단위 잔액 체인(잔액=직전+금액−수금) ②월계 공급가/부가세 대조 ③최종잔액 93,717,824.
출력: 품목마스터\\서울경금속_매입상세_2026상반기.xlsx (상세 + 품번 카탈로그) + JSON(적재용)."""
import pdfplumber, sys, re, json
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
BASE = r'Z:\Designs\123\서울경금속_매입_동산'
# (파일, 해당 파일 전기이월 — 명세 실측, 대상 월). 파일 간 이월 점프는 지연기표 지급분.
# ⚠️날짜·일계·월계는 PDF 텍스트 레이어에 없음(그래픽) → 파일1의 12월/1월 경계는
#   잔액 체인이 41,725,657(2025말 확정치)에 도달하는 12/31 현금 수금 직후로 판정. 월계 assert가 게이트.
FILES = [('1.동산기획 (2).pdf', 44944091, '2026-01'), ('2.동산기획 (3).pdf', 46198884, '2026-02'),
         ('3.동산기획 (3).pdf', 60637682, '2026-03'), ('4.동산기획 (2).pdf', 45730916, '2026-04'),
         ('5.동산기획.pdf', 54858496, '2026-05'), ('6.동산기획 (1).pdf', 66419386, '2026-06')]
YEAR_END_BAL = 41725657
XLSX = r'C:\Users\user\dongsan_mes\품목마스터\서울경금속_매입상세_2026상반기.xlsx'
JSON_OUT = r'C:\Users\user\dongsan_mes\docs\dongsan-import\load\sk_lines.json'

# 월계 정답 (명세 실측 — 공급가, 부가세)
MONTH_CHECK = {'2026-01': (16725000, 1672500), '2026-02': (13126180, 1312618),
               '2026-03': (28447380, 2844738), '2026-04': (21423980, 2142398),
               '2026-05': (10509900, 1050990), '2026-06': (42998580, 4299858)}


def num(s):
    s = str(s or '').replace(',', '').strip()
    if not s:
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


# 품명이 그래픽 레이어라 텍스트 추출이 빈칸인 품번 26종 — 렌더→판독 확정분(2026-08-01).
# prod 백필 = load/sk_name_backfill.sql. 재추출 시 여기서 보충해야 빈칸이 재발하지 않는다.
NAME_FIX = {
    '1371': ('보조시트(75)SC-008/ULTRA', ''),
    '1528': ('LG시트-보조(전사지)LC2000', ''),
    '1942': ('LG-ONEWAY1370mmLM54000', ''),
    '2498': ('사각경관봉70mm알미늄몸통3010mm', ''),
    '2499': ('사각경관봉70mmPC커버3010mm', ''),
    '26105': ('LG시트-LC6770(이형지개선품)', ''),
    '26474': ('YESLED형광등20W(보급형)/박스검정', '1BOX-50EA'),
    '26787': ('예스후렉스-(씨트용)120폭-50M', ''),
    '26796': ('예스후렉스-(그레이비조명용)90폭-50M', ''),
    '26797': ('예스후렉스-(그레이비조명용)100폭-50M', ''),
    '26798': ('예스후렉스-(그레이비조명용)110폭-50M', ''),
    '26799': ('예스후렉스-(그레이비조명용)120폭-50M', ''),
    '26800': ('예스후렉스-(그레이비조명용)130폭-50M', ''),
    '26801': ('예스후렉스-(그레이비조명용)150폭-50M', ''),
    '26802': ('예스후렉스-(그레이비조명용)160폭-50M', ''),
    '26803': ('예스후렉스-(그레이비조명용)180폭-50M', ''),
    '26804': ('예스후렉스-(그레이비조명용)200폭-50M', ''),
    '26806': ('예스후렉스-(그레이비조명용)260폭-50M', ''),
    '26807': ('예스후렉스-(그레이비조명용)320폭-50M', ''),
    '26894': ('YESLED형광등20W-(전구색,웜)', ''),
    '26895': ('YESLED형광등30W-(양면용,120cm)', ''),
    '26896': ('YESLED형광등15W-(양면용,60cm)', ''),
    '26898': ('YESLED형광등7W-(전구색,웜-60cm)', ''),
    '27097': ('유니온(UNION)-SMPS(고급형)-300W', ''),
    '27102': ('예스-고무자석롤-0.8T1020폭*10M', ''),
    '27103': ('예스-고무자석롤-0.8T610폭*10M', ''),
}


lines = []          # 2026년 상세행
pays = []           # 수금행 (검증용)

for fn, opening, month in FILES:
    bal = opening
    in_target = (month != '2026-01')   # 파일1만 12월 구간으로 시작 → 41,725,657 도달 후 1월
    with pdfplumber.open(BASE + '\\' + fn) as pdf:
        for page in pdf.pages:
            tbl = page.extract_table()
            if not tbl:
                continue
            for row in tbl:
                if not row or len(row) < 15:
                    continue
                c = [str(x or '').strip() for x in row]
                seq, _, carry, pno, pname, spec, qty, price, sup, vat, amt, note, form, pay, remain = c[:15]
                if not seq.isdigit():
                    continue
                amt_v, pay_v, rem_v = num(amt), num(pay), num(remain)
                if rem_v == 0 and amt_v == 0 and pay_v == 0:
                    continue
                # 행 체인 검산
                bal2 = bal + amt_v - pay_v
                assert abs(bal2 - rem_v) < 1, f'{fn} 행{seq} 체인 {bal2:,.0f} != 잔액 {rem_v:,.0f} ({pname})'
                bal = bal2
                if pay_v > 0:
                    pays.append((month, pay_v))
                    if not in_target and abs(bal - YEAR_END_BAL) < 1:
                        in_target = True   # 12/31 현금 수금 직후 = 2025말 잔액 도달 → 이후 1월
                    continue
                if in_target:
                    if not pname.strip() and pno in NAME_FIX:
                        pname, spec = NAME_FIX[pno][0], spec or NAME_FIX[pno][1]
                    lines.append(dict(d=month, pno=pno, name=pname, spec=spec,
                                      qty=num(qty), price=num(price), sup=num(sup), vat=num(vat), amt=amt_v))

# 검산 ②③
by_m = defaultdict(lambda: [0, 0])
for L in lines:
    k = L['d'][:7]
    by_m[k][0] += L['sup']
    by_m[k][1] += L['vat']
for k, (s, v) in MONTH_CHECK.items():
    gs, gv = by_m[k]
    assert round(gs) == s and round(gv) == v, f'{k} 월계 {gs:,.0f}/{gv:,.0f} != {s:,}/{v:,}'
assert abs(bal - 93717824) < 1, f'최종잔액 {bal:,.0f} != 93,717,824'
print(f'검산 3중 통과: 상세 {len(lines)}라인 · 수금 {len(pays)}건 {sum(p[1] for p in pays):,.0f} · 최종잔액 {bal:,.0f}')

# 품번 카탈로그
cat = {}
for L in lines:
    k = L['pno']
    if k not in cat:
        cat[k] = dict(pno=k, name=L['name'], spec=L['spec'], n=0, qty=0, sup=0, last_price=0)
    cat[k]['n'] += 1
    cat[k]['qty'] += L['qty']
    cat[k]['sup'] += L['sup']
    cat[k]['last_price'] = L['price']
    if len(L['name']) > len(cat[k]['name']):
        cat[k]['name'] = L['name']
print(f'품번 카탈로그: {len(cat)}종')

json.dump(dict(lines=lines, catalog=list(cat.values())), open(JSON_OUT, 'w', encoding='utf-8'), ensure_ascii=False)

# 엑셀
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hf = PatternFill('solid', fgColor='1F4E78')
hfont = Font(bold=True, color='FFFFFF')
thin = Border(*[Side(style='thin', color='D9D9D9')] * 4)

ws = wb.active
ws.title = '상세 라인'
ws.append(['일자', '품번', '품명', '규격', '수량', '단가', '공급가', '부가세', '금액(VAT포함)'])
for L in lines:
    ws.append([L['d'], L['pno'], L['name'], L['spec'], L['qty'], L['price'], L['sup'], L['vat'], L['amt']])
ws2 = wb.create_sheet('품번 카탈로그')
ws2.append(['품번', '품명', '규격', '라인수', '누계수량', '누계공급가', '최근단가'])
for v in sorted(cat.values(), key=lambda x: -x['sup']):
    ws2.append([v['pno'], v['name'], v['spec'], v['n'], v['qty'], v['sup'], v['last_price']])
for w in (ws, ws2):
    for j in range(1, w.max_column + 1):
        cell = w.cell(1, j)
        cell.fill, cell.font = hf, hfont
        w.column_dimensions[get_column_letter(j)].width = [11, 8, 30, 12, 9, 11, 12, 11, 13][j - 1] if j <= 9 else 12
    for row in w.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin
            if isinstance(cell.value, (int, float)) and cell.column >= 5:
                cell.number_format = '#,##0'
    w.freeze_panes = 'A2'
    w.auto_filter.ref = w.dimensions
wb.save(XLSX)
print('엑셀 저장:', XLSX)
