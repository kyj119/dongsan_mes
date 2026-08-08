# -*- coding: utf-8 -*-
"""이카운트 선명 품목 마스터 ↔ MES 갭 워크시트 (2026-08-08, 용준님 요청).

이카운트 942품목 중 **MES 연결 이력이 없는 344개**에 대해, MES 후보 코드를 채워 엑셀로 낸다.
용준님이 보고 「기존 코드로 붙일지 / 새로 만들지 / 안 쓸지」를 판단하는 게 목적이다.

★ 왜 일괄 등록을 안 하고 이 표를 만드는가
  344개 표본을 보면 **대부분 표기만 다를 뿐 MES 에 이미 있다**
  (`LED 3구(KPL)`=`SGM-LED3-KPL` · `돌출간판`=`SIGN-PRT` · `무광 솔벤 PVC(켈)-50m 127폭`=`KEL-SV50-127`).
  일괄 등록하면 **까치발 이중등록·황금봉 혼재를 대규모로 재생산**한다 —
  그 둘을 되돌리는 데 든 품이 정확히 이런 중복 때문이었다.

★ 점수는 **추천이지 판정이 아니다.** 자동 채택 임계값을 두지 않는다.
  매입은 재고가 움직여서 오연결이 나중에 흔적을 안 남긴다(선명 7월 이관에서 세운 원칙).
  그래서 이 표는 **사람이 채우는 열**(`★확정코드`)을 비워 둔 채로 나간다.

점수 구성 (0~1)
  토큰 자카드 0.5  +  폭/치수 일치 0.2  +  두께 일치 0.15  +  색상 일치 0.15
  - 폭·두께·색은 **불일치 시 감점**한다. 원단은 폭이 곧 단가고, 포맥스/아크릴은 두께가 곧 단가라
    「이름은 비슷한데 폭이 다른」 후보가 1위로 올라오면 원가가 조용히 틀린다.

열 구성
  이카운트코드 · 품목명 · 규격 · 품목구분 · 그룹 · 출고가 · 입고가 · 구매처
  → 후보1~3(코드·품목명·점수·사유) · **자동판단** · **★확정코드(사람 기입)**

자동판단
  「기존연결 유력」 0.75+ · 「검토」 0.45~0.75 · 「신설 후보」 0.45 미만
  거래 실적(출고가·입고가) 유무를 `우선순위` 열로 함께 낸다 — 실적 없는 건 등록 자체가 불필요할 수 있다.

사용: python docs/dongsan-import/gen_ecount_item_gap.py <ecount_missing.json> <mes_items.json>
산출: docs/analysis/2026-08-08-ecount-item-gap.xlsx
"""
import json, sys, os, re, warnings
from collections import Counter
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

MISS_JSON, MES_JSON = sys.argv[1], sys.argv[2]
OUT = os.path.join(r'C:\Users\user\dongsan_mes', 'docs', 'analysis', '2026-08-08-ecount-item-gap.xlsx')

miss = json.load(open(MISS_JSON, encoding='utf-8'))
mes = json.load(open(MES_JSON, encoding='utf-8'))

# ── 정규화: 매칭을 방해하는 표기 차이를 걷어낸다
NOISE = re.compile(r'[()\[\]/·\-_,\.]|30m|50m|45m|61m|\d+g|ea|EA')
SPLIT = re.compile(r'[\s]+')
WIDTH = re.compile(r'(\d{2,3})\s*(?:폭|cm)')
THICK = re.compile(r'(\d+(?:\.\d+)?)\s*T')
SIZE = re.compile(r'(\d{2,4})\s*mm')
# ★ 색상 사전은 **긴 이름부터** 넣는다. dict 는 삽입 순서를 보존하므로
#   「진회색」이 「회색」보다 먼저 걸린다. 첫 판에서 이걸 안 지켜 진회색·연회색이 둘 다 'G' 로 뭉개졌고,
#   「채널바 옆마구리 300mm*100mm/진회색」이 연회색 코드(`SGM-CSM-R300-LG`)와 0.92 로 붙었다.
#   간판 자재는 색이 코드 접미(BK/WH/DG/LG/NB/LB/DP/RD/GN/YL/OR/BR/JD/FL)로 갈리는 축이라
#   색을 뭉개면 **다른 물건에 붙는다**.
COLOR = {
    '진회색': 'DG', '연회색': 'LG', '진청색': 'NB', '연청색': 'LB', '진핑크': 'DP', '연핑크': 'LP',
    '주황색': 'OR', '형광색': 'FL', '형광': 'FL', '밤색': 'BR', '옥색': 'JD', '녹색': 'GN',
    '적색': 'RD', '황색': 'YL', '금색': 'GD', '연두': 'YG',
    '검정': 'BK', '흑색': 'BK', '블랙': 'BK', '백색': 'WH', '흰': 'WH', '투명': 'C',
    '그레이': 'LG', '회색': 'LG', '은색': 'S', '웜': 'M',
}
# 관례상 동일 폭 (용준님 확인: 150≡152 · 122≡120)
WEQ = {150: 152, 152: 152, 122: 120, 120: 120}


def norm_tokens(*parts):
    t = NOISE.sub(' ', ' '.join(p for p in parts if p))
    return {w for w in SPLIT.split(t) if len(w) >= 2}


def bigrams(*parts):
    """문자 2-gram. **한국어 붙여쓰기 때문에 필요하다** — 「LG조명용시트」는 공백이 없어
    단어 토큰화로는 통째로 한 덩어리가 되고, MES 의 「조명용 시트」와 전혀 안 겹친다."""
    t = NOISE.sub('', ' '.join(p for p in parts if p)).replace(' ', '')
    return {t[i:i + 2] for i in range(len(t) - 1)}


def width_of(*parts):
    for p in parts:
        m = WIDTH.search(p or '')
        if m:
            w = int(m.group(1))
            return WEQ.get(w, w)
    return None


def thick_of(*parts):
    for p in parts:
        m = THICK.search(p or '')
        if m:
            return m.group(1).rstrip('0').rstrip('.')
    return None


def size_of(*parts):
    for p in parts:
        m = SIZE.search(p or '')
        if m:
            return int(m.group(1))
    return None


def color_of(*parts):
    s = ' '.join(p for p in parts if p)
    for k, v in COLOR.items():
        if k in s:
            return v
    return None


# 같은 이름에 **규격 있는 형제**가 있는지 미리 센다.
#   `SGM-TRB`(규격 없음)와 `SGM-TRB-R200-DG`(200mm/진회색)가 공존하는데, 이카운트가
#   「트러스바 200mm*7m/진회색」으로 오면 이름이 같은 규격없는 대표 코드가 1.00 으로 1위가 된다.
#   대표 코드는 대개 **제품**(출력물)이고 규격별이 **자재**라, 그대로 붙이면 물건이 바뀐다.
from collections import defaultdict as _dd
_spec_sibs = _dd(int)
for _m in mes:
    if str(_m['sp']).strip():
        _spec_sibs[_m['item_name'].strip()] += 1

MES_IDX = []
for m in mes:
    MES_IDX.append({
        'code': m['item_code'], 'name': m['item_name'], 'sp': m['sp'], 'cat': m['category'],
        'tok': norm_tokens(m['item_name'], m['sp']), 'bg': bigrams(m['item_name'], m['sp']),
        'w': width_of(m['sp'], m['item_name']),
        't': thick_of(m['item_name'], m['sp']), 'sz': size_of(m['sp'], m['item_name']),
        'c': color_of(m['item_name'], m['sp']),
        'bare': not str(m['sp']).strip() and _spec_sibs.get(m['item_name'].strip(), 0) > 0,
    })


def score(e, m):
    """0~1. 이름 유사도가 **단독으로 1.0 까지** 갈 수 있어야 한다 —
    「돌출간판」처럼 완전히 같은 이름이 상한 0.5 에 걸리면, 명백한 기존 품목이 「신설 후보」로 나와
    이 표가 오히려 중복 등록을 부추긴다(첫 판에서 실제로 그랬다).

    이름 = 단어 자카드와 문자 2-gram 자카드 중 **큰 쪽**. 한국어는 띄어쓰기가 불규칙해서
    둘 중 하나만 쓰면 반드시 한쪽 계열을 놓친다.
    폭·두께·색은 **가점보다 감점이 크게** — 원단은 폭이, 포맥스/아크릴은 두께가 곧 단가라
    「이름은 비슷한데 폭이 다른」 후보가 1위로 오르면 원가가 조용히 틀린다."""
    if not e['tok'] and not e['bg']:
        return 0.0, ''
    jt = len(e['tok'] & m['tok']) / len(e['tok'] | m['tok']) if (e['tok'] | m['tok']) else 0.0
    jb = len(e['bg'] & m['bg']) / len(e['bg'] | m['bg']) if (e['bg'] | m['bg']) else 0.0
    base = max(jt, jb)
    s = base
    why = ['이름 %.2f(단어 %.2f·자소 %.2f)' % (base, jt, jb)]
    hits = misses = 0
    for key, up, dn, label in (('w', 0.10, 0.35, '폭'), ('t', 0.08, 0.30, '두께'), ('c', 0.06, 0.20, '색')):
        ev, mv = e[key], m[key]
        if ev is None or mv is None:
            continue
        if ev == mv:
            s += up
            hits += 1
            why.append('%s일치' % label)
        else:
            s -= dn
            misses += 1
            why.append('%s불일치(%s≠%s)' % (label, ev, mv))
    # 규격 없는 대표 코드인데 **이카운트 쪽엔 규격이 있다** → 규격별 형제가 정본이다
    if m['bare'] and (e['w'] or e['t'] or e['sz'] or e['c']):
        s -= 0.30
        why.append('대표코드(규격없음)·형제有')
    if e['sz'] and m['sz']:
        # 치수 불일치는 폭과 같은 급으로 깎는다 — 「보강대 150mm」가 `SGM-TRF-R250`(250mm)에
        #   0.85 로 붙어 「기존연결 유력」으로 나왔다. MES 에 150 이 없으면 **신설 후보가 정답**이다.
        if e['sz'] == m['sz']:
            s += 0.08
            hits += 1
        else:
            s -= 0.35
            misses += 1
        why.append('치수' + ('일치' if e['sz'] == m['sz'] else '불일치'))
    # ★ 속성이 **둘 이상 일치하고 하나도 어긋나지 않으면** 이름 점수가 낮아도 사실상 확정이다.
    #   「채널용 입체바 진회색/100mm*3m」과 「입체바 1.0T 100mm 진회색」은 글자가 많이 달라
    #   이름 0.60 에 그치는데, **색과 치수가 둘 다 맞는데 다른 물건일 확률**은 낮다.
    #   이름은 표기 습관이고 색·치수·폭·두께는 **물건을 가르는 축**이다 — 후자를 더 믿는다.
    if hits >= 2 and misses == 0:
        s += 0.15
        why.append('속성 %d개 일치·불일치 없음' % hits)
    return max(0.0, min(1.0, s)), ' · '.join(why)


rows, stat = [], Counter()
for x in miss:
    e = {'tok': norm_tokens(x['nm'], x['sp']), 'bg': bigrams(x['nm'], x['sp']),
         'w': width_of(x['sp'], x['nm']),
         't': thick_of(x['nm'], x['sp']), 'sz': size_of(x['sp'], x['nm']),
         'c': color_of(x['nm'], x['sp'])}
    cands = sorted(((score(e, m), m) for m in MES_IDX), key=lambda p: -p[0][0])[:3]
    top = cands[0][0][0] if cands else 0.0
    verdict = '기존연결 유력' if top >= 0.75 else ('검토' if top >= 0.45 else '신설 후보')
    stat[verdict] += 1
    has_price = bool(x['out'] or x['inp'])
    row = [x['code'], x['nm'], x['sp'], x['kind'], x['g1'],
           x['out'] or '', x['inp'] or '', x['sup'],
           '실적단가 있음' if has_price else '단가 없음']
    for (sc, why), m in cands:
        row += [m['code'], m['name'] + ((' / ' + m['sp']) if m['sp'] else ''), round(sc, 2)]
    while len(row) < 9 + 9:
        row.append('')
    row += [cands[0][0][1] if cands else '', verdict, '']
    rows.append(row)

rows.sort(key=lambda r: (r[-2] != '신설 후보', -float(r[11] or 0)))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '이카운트 품목 갭'
HDR = ['이카운트코드', '품목명', '규격', '구분', '그룹', '출고가', '입고가', '구매처', '우선순위',
       '후보1코드', '후보1명', '점수1', '후보2코드', '후보2명', '점수2', '후보3코드', '후보3명', '점수3',
       '판정근거', '자동판단', '★확정코드(기입)']
ws.append(HDR)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for r in rows:
    ws.append(r)
FILL = {'신설 후보': 'FFF2CC', '검토': 'FCE4D6', '기존연결 유력': 'E2EFDA'}
for i in range(2, ws.max_row + 1):
    v = ws.cell(i, 20).value
    if v in FILL:
        for j in range(1, len(HDR) + 1):
            ws.cell(i, j).fill = PatternFill('solid', fgColor=FILL[v])
for col, w in zip('ABCDEFGHIJKLMNOPQRSTU',
                  [13, 34, 16, 9, 8, 10, 10, 16, 12, 15, 30, 7, 15, 30, 7, 15, 30, 7, 40, 13, 18]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'C2'
ws.auto_filter.ref = ws.dimensions
# 엑셀이 열려 있으면 잠겨서 저장이 막힌다 — 조용히 실패시키지 않고 **옆 파일로 낸다**.
#   (검토 중에 재생성하는 게 정상 흐름이라, 닫으라고 요구하는 것보다 이쪽이 낫다.)
try:
    wb.save(OUT)
except PermissionError:
    OUT = OUT.replace('.xlsx', '-v2.xlsx')
    wb.save(OUT)
    print('⚠️ 원본이 열려 있어 %s 로 저장' % os.path.basename(OUT))

print('%d행 → %s' % (len(rows), OUT))
print('자동판단:', dict(stat))
print('실적단가 있는 것 %d' % sum(1 for x in miss if x['out'] or x['inp']))
