#!/usr/bin/env python3
"""원본 엑셀과 변환 엑셀 간 데이터 무결성 검증 스크립트."""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

ORIG_FILE = Path(__file__).parent.parent / "거래처정보.xlsx"
CONV_FILE = Path(__file__).parent.parent / "거래처정보_MES변환.xlsx"


def safe(val):
    if val is None:
        return ""
    return str(val).strip()


def load_original():
    """원본 엑셀 로드."""
    wb = load_workbook(ORIG_FILE, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and safe(row[0]) == "거래처코드":
                header_found = True
            continue
        code = safe(row[0]) if row and len(row) > 0 else ""
        if not re.match(r"^\d{3}-\d{2}-\d{5}$", code):
            continue
        data[code] = {
            "name": safe(row[1]),
            "rep": safe(row[2]),
            "btype": safe(row[3]),
            "bitem": safe(row[4]),
            "fax": safe(row[5]),
            "email": safe(row[6]),
            "addr1": safe(row[7]),
            "addr2": safe(row[8]) if len(row) > 8 else "",
            "phone": safe(row[10]) if len(row) > 10 else "",
            "mobile": safe(row[11]) if len(row) > 11 else "",
            "search": safe(row[12]) if len(row) > 12 else "",
            "invoice": safe(row[13]) if len(row) > 13 else "",
            "active": safe(row[14]) if len(row) > 14 else "",
            "transfer": safe(row[15]) if len(row) > 15 else "",
        }
    wb.close()
    return data


def load_converted():
    """변환 엑셀 로드."""
    wb = load_workbook(CONV_FILE, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            continue
        code = safe(row[0]) if row and len(row) > 0 else ""
        if not code:
            continue
        data[code] = {
            "client_code": code,
            "name": safe(row[1]),
            "brn": safe(row[2]),
            "rep": safe(row[3]),
            "btype": safe(row[4]),
            "bitem": safe(row[5]),
            "phone": safe(row[6]),
            "mobile": safe(row[7]),
            "fax": safe(row[8]),
            "email": safe(row[9]),
            "address": safe(row[10]),
            "address_detail": safe(row[11]),
            "delivery_method": safe(row[12]),
            "delivery_address": safe(row[13]),
            "invoice_type": safe(row[14]),
            "invoice_raw": safe(row[15]),
            "search": safe(row[16]),
            "transfer": safe(row[17]),
            "is_active": safe(row[18]),
            "addr1_orig": safe(row[19]),
            "addr2_orig": safe(row[20]),
        }
    wb.close()
    return data


INVOICE_MAP = {
    "일별": "PER_ORDER",
    "월별": "MONTHLY",
    "주별": "MONTHLY",
    "15일": "MONTHLY",
    "미분류": "UNDECIDED",
    "카드결재": "CARD",
    "카드결제": "CARD",
    "타발행": "ISSUED_BY_OTHER",
}


def main():
    print("원본 로드 중...")
    original = load_original()
    print(f"원본: {len(original)}건")

    print("변환 파일 로드 중...")
    converted = load_converted()
    print(f"변환: {len(converted)}건")

    errors = []
    warnings = []

    # ── 1. 건수 비교 ─────────────────────────────
    if len(original) != len(converted):
        errors.append(f"건수 불일치! 원본={len(original)} vs 변환={len(converted)}")
        missing = set(original.keys()) - set(converted.keys())
        extra = set(converted.keys()) - set(original.keys())
        if missing:
            errors.append(f"  변환에서 누락: {len(missing)}건")
            for c in sorted(missing)[:5]:
                errors.append(f"    {c} - {original[c]['name']}")
        if extra:
            errors.append(f"  변환에만 존재: {len(extra)}건")

    # ── 2. 직접 복사 필드 대조 ────────────────────
    fields_to_check = [
        ("name", "name"),
        ("rep", "rep"),
        ("btype", "btype"),
        ("bitem", "bitem"),
        ("fax", "fax"),
        ("email", "email"),
        ("phone", "phone"),
        ("mobile", "mobile"),
        ("search", "search"),
        ("transfer", "transfer"),
    ]

    print("\n[직접 복사 필드 대조]")
    common_codes = sorted(set(original.keys()) & set(converted.keys()))

    for field_orig, field_conv in fields_to_check:
        mismatches = []
        for code in common_codes:
            ov = original[code][field_orig]
            cv = converted[code][field_conv]
            if ov != cv:
                mismatches.append((code, original[code]["name"], ov, cv))
        status = "OK" if not mismatches else f"{len(mismatches)}건 불일치"
        print(f"  {field_orig:12s}: {status}")
        for code, name, ov, cv in mismatches[:3]:
            print(f"    예시: {name} 원본=[{ov[:30]}] 변환=[{cv[:30]}]")
        if mismatches:
            errors.append(f"{field_orig} 불일치 {len(mismatches)}건")

    # ── 3. 주소 원본 보존 확인 ────────────────────
    print("\n[주소 원본 보존]")
    addr1_mismatch = 0
    addr2_mismatch = 0
    for code in common_codes:
        if original[code]["addr1"] != converted[code]["addr1_orig"]:
            addr1_mismatch += 1
            if addr1_mismatch <= 2:
                warnings.append(
                    f"addr1 원본 변경: {code} "
                    f"[{original[code]['addr1'][:35]}] vs "
                    f"[{converted[code]['addr1_orig'][:35]}]"
                )
        if original[code]["addr2"] != converted[code]["addr2_orig"]:
            addr2_mismatch += 1
            if addr2_mismatch <= 2:
                warnings.append(
                    f"addr2 원본 변경: {code} "
                    f"[{original[code]['addr2'][:35]}] vs "
                    f"[{converted[code]['addr2_orig'][:35]}]"
                )
    print(f"  addr1 원본 보존: {'OK' if addr1_mismatch == 0 else f'{addr1_mismatch}건 불일치'}")
    print(f"  addr2 원본 보존: {'OK' if addr2_mismatch == 0 else f'{addr2_mismatch}건 불일치'}")

    # ── 4. 주소 파싱 내용 유실 확인 ───────────────
    print("\n[주소 파싱 내용 유실 확인]")
    content_lost = 0
    lost_examples = []
    for code in common_codes:
        orig_addr = original[code]["addr1"]
        if not orig_addr:
            continue
        parsed = converted[code]["address"] + " " + converted[code]["address_detail"]
        # 원본의 모든 숫자(번지)가 파싱 결과에 포함되는지
        orig_nums = set(re.findall(r"\d+", orig_addr))
        parsed_nums = set(re.findall(r"\d+", parsed))
        lost = orig_nums - parsed_nums
        if lost:
            content_lost += 1
            if len(lost_examples) < 5:
                lost_examples.append((
                    original[code]["name"],
                    orig_addr[:55],
                    parsed.strip()[:55],
                    lost,
                ))
    print(f"  번지 유실: {'OK' if content_lost == 0 else f'{content_lost}건'}")
    for name, orig, parsed, lost in lost_examples:
        print(f"    {name}:")
        print(f"      원본:  {orig}")
        print(f"      파싱:  {parsed}")
        print(f"      유실:  {lost}")
    if content_lost:
        warnings.append(f"주소 번지 유실 {content_lost}건")

    # 원본 한글 단어 유실도 확인
    print("\n[주소 파싱 한글 키워드 유실 확인]")
    keyword_lost = 0
    keyword_examples = []
    for code in common_codes:
        orig_addr = original[code]["addr1"]
        if not orig_addr:
            continue
        parsed = converted[code]["address"] + " " + converted[code]["address_detail"]
        # 원본의 핵심 한글 단어 (2자 이상)
        orig_words = set(re.findall(r"[가-힣]{2,}", orig_addr))
        parsed_words = set(re.findall(r"[가-힣]{2,}", parsed))
        lost = orig_words - parsed_words
        if lost:
            keyword_lost += 1
            if len(keyword_examples) < 5:
                keyword_examples.append((original[code]["name"], lost))
    print(f"  한글 키워드 유실: {'OK' if keyword_lost == 0 else f'{keyword_lost}건'}")
    for name, lost in keyword_examples:
        print(f"    {name}: {lost}")
    if keyword_lost:
        warnings.append(f"주소 한글 키워드 유실 {keyword_lost}건")

    # ── 5. invoice_type 매핑 검증 ─────────────────
    print("\n[invoice_type 매핑]")
    inv_mismatch = 0
    inv_examples = []
    for code in common_codes:
        raw = original[code]["invoice"]
        expected = INVOICE_MAP.get(raw, "UNDECIDED")
        actual = converted[code]["invoice_type"]
        if expected != actual:
            inv_mismatch += 1
            if len(inv_examples) < 5:
                inv_examples.append((original[code]["name"], raw, expected, actual))
    print(f"  매핑 정확도: {'OK' if inv_mismatch == 0 else f'{inv_mismatch}건 오류'}")
    for name, raw, exp, act in inv_examples:
        print(f"    {name}: 원본=[{raw}] 기대=[{exp}] 실제=[{act}]")
    if inv_mismatch:
        errors.append(f"invoice_type 매핑 오류 {inv_mismatch}건")

    # ── 6. client_code == BRN 확인 ────────────────
    print("\n[client_code == business_registration_number]")
    brn_mismatch = sum(
        1 for c in converted.values() if c["client_code"] != c["brn"]
    )
    print(f"  일치: {'OK' if brn_mismatch == 0 else f'{brn_mismatch}건 불일치'}")

    # ── 7. 중복 코드 확인 ────────────────────────
    print("\n[사업자번호 중복]")
    from collections import Counter
    code_counts = Counter(converted.keys())
    dupes = {k: v for k, v in code_counts.items() if v > 1}
    print(f"  중복: {'없음' if not dupes else f'{len(dupes)}건'}")
    for code, cnt in list(dupes.items())[:5]:
        print(f"    {code}: {cnt}회")

    # ── 최종 결과 ─────────────────────────────────
    print("\n" + "=" * 60)
    if not errors and not warnings:
        print("  >>> 데이터 무결성 검증 통과 - 유실/뒤틀림 없음 <<<")
    else:
        if errors:
            print("  *** 오류 ***")
            for e in errors:
                print(f"    {e}")
        if warnings:
            print("  *** 경고 (확인 필요) ***")
            for w in warnings:
                print(f"    {w}")
    print("=" * 60)


if __name__ == "__main__":
    main()
