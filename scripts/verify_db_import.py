#!/usr/bin/env python3
"""원본 엑셀 vs MES DB 데이터 대조 검증."""

import sys
import re
import json
import subprocess
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

ORIG_FILE = Path(__file__).parent.parent / "거래처정보.xlsx"
PROJECT_DIR = Path(__file__).parent.parent

INVOICE_MAP = {
    "일별": "PER_ORDER",
    "월별": "MONTHLY",
    "주별": "MONTHLY",
    "15일": "MONTHLY",
    "미분류": "UNDECIDED",
    "카드결재": "CARD",
    "카드결제": "CARD",
    "타발행": "ISSUED_BY_OTHER",
}


def safe(val):
    if val is None:
        return ""
    return str(val).strip()


def query_db(sql):
    """wrangler d1 execute로 DB 조회."""
    result = subprocess.run(
        ["npx", "wrangler", "d1", "execute", "DB", "--local", "--command", sql, "--json"],
        capture_output=True, text=True, cwd=PROJECT_DIR, encoding="utf-8"
    )
    try:
        data = json.loads(result.stdout)
        return data[0]["results"] if data else []
    except (json.JSONDecodeError, KeyError, IndexError):
        print(f"DB 쿼리 실패: {result.stderr[:200]}")
        return []


def main():
    # 원본 로드
    print("원본 엑셀 로드 중...")
    wb = load_workbook(ORIG_FILE, read_only=True, data_only=True)
    ws = wb.active
    original = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and safe(row[0]) == "거래처코드":
                header_found = True
            continue
        code = safe(row[0]) if row and len(row) > 0 else ""
        if not re.match(r"^\d{3}-\d{2}-\d{5}$", code):
            continue
        original[code] = {
            "name": safe(row[1]),
            "rep": safe(row[2]),
            "btype": safe(row[3]),
            "bitem": safe(row[4]),
            "fax": safe(row[5]),
            "email": safe(row[6]),
            "addr1": safe(row[7]),
            "addr2": safe(row[8]) if len(row) > 8 else "",
            "phone": safe(row[10]) if len(row) > 10 else "",
            "mobile": safe(row[11]) if len(row) > 11 else "",
            "search": safe(row[12]) if len(row) > 12 else "",
            "invoice": safe(row[13]) if len(row) > 13 else "",
        }
    wb.close()
    print(f"원본: {len(original)}건")

    # DB에서 랜덤 10건 + 특정 케이스 추출
    print("\nDB 데이터 조회 중...")

    # 랜덤 10건
    random_rows = query_db(
        "SELECT client_code, client_name, representative, business_type, business_item, "
        "phone, mobile, fax, email, address, address_detail, search_keywords, "
        "business_registration_number, delivery_method, delivery_address, invoice_method, transfer_info "
        "FROM clients ORDER BY RANDOM() LIMIT 10"
    )

    # 기존 2건 (선명커뮤니케이션, 계룡기획인쇄)
    existing_rows = query_db(
        "SELECT client_code, client_name, representative, business_type, business_item, "
        "phone, mobile, fax, email, address, address_detail, search_keywords, "
        "business_registration_number, delivery_method, delivery_address, invoice_method, transfer_info "
        "FROM clients WHERE client_name IN ('선명커뮤니케이션', '계룡기획인쇄')"
    )

    # FREIGHT 배송 샘플
    freight_rows = query_db(
        "SELECT client_code, client_name, delivery_method, delivery_address "
        "FROM clients WHERE delivery_method = 'FREIGHT' LIMIT 5"
    )

    # DIRECT 배송 샘플
    direct_rows = query_db(
        "SELECT client_code, client_name, delivery_method, delivery_address "
        "FROM clients WHERE delivery_method = 'DIRECT' LIMIT 5"
    )

    # invoice_method 분포
    invoice_dist = query_db(
        "SELECT invoice_method, COUNT(*) as cnt FROM clients GROUP BY invoice_method ORDER BY cnt DESC"
    )

    # transfer_info 확인
    transfer_check = query_db(
        "SELECT COUNT(*) as cnt FROM clients WHERE transfer_info IS NOT NULL AND transfer_info != '' AND transfer_info != '등록'"
    )

    # ── 검증 출력 ─────────────────────────────────

    print("\n" + "=" * 70)
    print("  원본 엑셀 vs DB 랜덤 샘플 대조 (10건)")
    print("=" * 70)

    errors = 0
    for db_row in random_rows:
        code = db_row["client_code"]
        if code not in original:
            print(f"\n  [!] {code} - DB에 있지만 원본에 없음")
            errors += 1
            continue

        orig = original[code]
        mismatches = []

        # 직접 비교 필드
        checks = [
            ("거래처명", orig["name"], db_row["client_name"]),
            ("대표자", orig["rep"], db_row["representative"] or ""),
            ("업태", orig["btype"], db_row["business_type"] or ""),
            ("종목", orig["bitem"], db_row["business_item"] or ""),
            ("전화", orig["phone"], db_row["phone"] or ""),
            ("모바일", orig["mobile"], db_row["mobile"] or ""),
            ("FAX", orig["fax"], db_row["fax"] or ""),
            ("이메일", orig["email"], db_row["email"] or ""),
            ("사업자번호", code, db_row["business_registration_number"] or ""),
        ]

        for label, ov, dv in checks:
            if ov != dv:
                mismatches.append(f"{label}: 원본=[{ov[:25]}] DB=[{dv[:25]}]")

        # invoice_method 매핑 확인
        expected_inv = INVOICE_MAP.get(orig["invoice"], "UNDECIDED")
        actual_inv = db_row["invoice_method"] or "PER_ORDER"
        if expected_inv != actual_inv:
            mismatches.append(f"계산서: 원본=[{orig['invoice']}→{expected_inv}] DB=[{actual_inv}]")

        status = "OK" if not mismatches else f"{len(mismatches)}건 불일치"
        print(f"\n  {code} ({db_row['client_name']}): {status}")
        for m in mismatches:
            print(f"    - {m}")
            errors += 1

    print("\n" + "=" * 70)
    print("  기존 등록 거래처 확인")
    print("=" * 70)
    for row in existing_rows:
        print(f"  {row['client_code']} ({row['client_name']})")
        print(f"    client_code=사업자번호: {'OK' if row['client_code'] == row['business_registration_number'] else 'FAIL'}")
        print(f"    delivery: {row['delivery_method']} / {row['delivery_address'] or '-'}")

    print("\n" + "=" * 70)
    print("  배송정보 샘플")
    print("=" * 70)
    print("  [FREIGHT]")
    for row in freight_rows:
        print(f"    {row['client_name']}: {row['delivery_address']}")
    print("  [DIRECT]")
    for row in direct_rows:
        print(f"    {row['client_name']}: {row['delivery_address'][:50] if row['delivery_address'] else '-'}")

    print("\n" + "=" * 70)
    print("  invoice_method 분포")
    print("=" * 70)
    for row in invoice_dist:
        print(f"    {row['invoice_method'] or '(NULL)':20s}: {row['cnt']}건")

    print("\n" + "=" * 70)
    print("  이체정보 확인")
    print("=" * 70)
    cnt = transfer_check[0]["cnt"] if transfer_check else 0
    print(f"    '등록' 아닌 실제 이체정보: {cnt}건")

    print("\n" + "=" * 70)
    if errors == 0:
        print("  >>> 데이터 검증 통과 <<<")
    else:
        print(f"  >>> {errors}건 불일치 발견 — 확인 필요 <<<")
    print("=" * 70)


if __name__ == "__main__":
    main()
