#!/usr/bin/env python3
"""DB 주소 파싱 오류 + CARD 매핑 일괄 수정 스크립트.

원본 엑셀에서 주소를 다시 읽어 DB를 보정한다.
수정 내용:
1. CARD 80건: 원본 '발행X,카드결재' → invoice_method = 'CARD'
2. (XX동) 상세주소 오류: 법정동 괄호를 기본주소로 복원
3. 층/호 미분리: 기본주소에 남은 층/호를 상세주소로 이동
"""

import sys
import re
import json
import subprocess
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PROJECT_DIR = Path(__file__).parent.parent
ORIG_FILE = PROJECT_DIR / "거래처정보.xlsx"


def run_sql(sql):
    """wrangler d1 execute 실행."""
    result = subprocess.run(
        ["npx.cmd", "wrangler", "d1", "execute", "DB", "--local", "--command", sql],
        capture_output=True, text=True, cwd=PROJECT_DIR, encoding="utf-8", errors="replace"
    )
    if result.returncode != 0:
        print(f"  SQL 오류: {result.stderr[:200]}")
        return False
    return True


def run_sql_json(sql):
    """wrangler d1 execute --json 실행."""
    result = subprocess.run(
        ["npx.cmd", "wrangler", "d1", "execute", "DB", "--local", "--json", "--command", sql],
        capture_output=True, text=True, cwd=PROJECT_DIR, encoding="utf-8", errors="replace"
    )
    try:
        data = json.loads(result.stdout)
        return data[0]["results"] if data else []
    except (json.JSONDecodeError, KeyError, IndexError):
        return []


def main():
    # ── 1. CARD 80건 수정 ─────────────────────────────
    print("=" * 60)
    print("  1. CARD 매핑 수정 (발행X,카드결재 → CARD)")
    print("=" * 60)

    # 원본에서 카드결재 사업자번호 추출
    wb = load_workbook(ORIG_FILE, read_only=True, data_only=True)
    ws = wb.active
    card_codes = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0] or "").strip() == "거래처코드":
                header_found = True
            continue
        code = str(row[0] or "").strip()
        if not re.match(r"^\d{3}-\d{2}-\d{5}$", code):
            continue
        inv = str(row[13] or "").strip() if len(row) > 13 else ""
        if "카드" in inv:
            card_codes.append(code)
    wb.close()

    print(f"  원본에서 카드결재 거래처: {len(card_codes)}건")

    # 배치로 UPDATE (50건씩)
    updated = 0
    for i in range(0, len(card_codes), 50):
        batch = card_codes[i:i+50]
        codes_str = ",".join(f"'{c}'" for c in batch)
        sql = f"UPDATE clients SET invoice_method = 'CARD' WHERE client_code IN ({codes_str})"
        if run_sql(sql):
            updated += len(batch)
    print(f"  CARD 업데이트: {updated}건")

    # ── 2. (XX동) 상세주소 → 기본주소 복원 ──────────────
    print()
    print("=" * 60)
    print("  2. 법정동 괄호 복원 (상세 → 기본주소)")
    print("=" * 60)

    # Case A: address_detail이 "(XX동)" 만 있는 경우 → address에 합치고 detail 비움
    rows_a = run_sql_json(
        "SELECT id, address, address_detail FROM clients "
        "WHERE address_detail LIKE '(%)'  "
        "AND length(address_detail) <= 10 "
        "AND address_detail NOT LIKE '%호%' AND address_detail NOT LIKE '%층%'"
    )
    print(f"  Case A (법정동만): {len(rows_a)}건")
    for row in rows_a:
        new_addr = row["address"] + " " + row["address_detail"]
        sql = f"UPDATE clients SET address = '{new_addr.replace(chr(39), chr(39)+chr(39))}', address_detail = '' WHERE id = {row['id']}"
        run_sql(sql)

    # Case B: address_detail이 "(XX동) 3층" 같은 경우 → (동)은 address로, 나머지는 detail 유지
    rows_b = run_sql_json(
        "SELECT id, address, address_detail FROM clients "
        "WHERE address_detail LIKE '(%) %' "
        "AND address_detail NOT LIKE '(%호%' "
    )
    print(f"  Case B (법정동+상세): {len(rows_b)}건")
    for row in rows_b:
        detail = row["address_detail"]
        # (XX동) 부분 추출
        m = re.match(r"(\([^)]+\))\s*(.*)", detail)
        if m:
            dong_part = m.group(1)
            rest_part = m.group(2).strip()
            new_addr = row["address"] + " " + dong_part
            esc = lambda s: s.replace("'", "''")
            sql = f"UPDATE clients SET address = '{esc(new_addr)}', address_detail = '{esc(rest_part)}' WHERE id = {row['id']}"
            run_sql(sql)

    # ── 3. 층/호 미분리 보정 ─────────────────────────────
    print()
    print("=" * 60)
    print("  3. 층/호 미분리 보정")
    print("=" * 60)

    # 주소 끝에 "N층", "N호", "N동N호" 패턴이 있는데 상세주소가 비어있는 경우
    rows_c = run_sql_json(
        "SELECT id, client_name, address, address_detail FROM clients "
        "WHERE (address_detail IS NULL OR address_detail = '') "
        "AND address != '' "
    )

    fixed_count = 0
    for row in rows_c:
        addr = row["address"]
        # 패턴: 주소 끝부분에서 층/호 분리
        # "... 1층", "... 401호", "... 1동 501호", "... 빌딩 201호", "... B동 3층"
        # 하지만 도로명에 포함된 호/층은 제외해야 함 (천호대로, 호국로 등)

        # 안전한 패턴: 공백 후 숫자+층/호로 끝나는 경우
        m = re.search(r"\s(\d+[-\d]*층[\d호]*)$", addr)
        if not m:
            m = re.search(r"\s(\d+[-\d]*호)$", addr)
        if not m:
            # "빌딩명 N층" 패턴
            m = re.search(r"\s(\S+\s+\d+층\S*)$", addr)
            if m and any(kw in m.group(1) for kw in ["로", "길", "대로"]):
                m = None  # 도로명이면 무시
        if not m:
            # "N동 N호" 패턴 (아파트)
            m = re.search(r"\s(\d+동\s*\d+호)$", addr)

        if m:
            detail_part = m.group(1).strip()
            base_part = addr[:m.start()].strip()
            # 도로명/지번이 남아있는지 확인 (너무 짧으면 무시)
            if len(base_part) > 10:
                esc = lambda s: s.replace("'", "''")
                sql = f"UPDATE clients SET address = '{esc(base_part)}', address_detail = '{esc(detail_part)}' WHERE id = {row['id']}"
                if run_sql(sql):
                    fixed_count += 1

    print(f"  층/호 분리 보정: {fixed_count}건")

    # ── 결과 확인 ─────────────────────────────────────
    print()
    print("=" * 60)
    print("  수정 결과 확인")
    print("=" * 60)

    inv_dist = run_sql_json(
        "SELECT invoice_method, COUNT(*) as cnt FROM clients GROUP BY invoice_method ORDER BY cnt DESC"
    )
    print("  [invoice_method 분포]")
    for r in inv_dist:
        print(f"    {r['invoice_method'] or '(NULL)':20s}: {r['cnt']}건")

    detail_stats = run_sql_json(
        "SELECT "
        "SUM(CASE WHEN address_detail LIKE '(%동)' OR address_detail LIKE '(%가)' OR address_detail LIKE '(%리)' THEN 1 ELSE 0 END) as dong_only, "
        "SUM(CASE WHEN address_detail LIKE '(%) %' THEN 1 ELSE 0 END) as dong_plus, "
        "SUM(CASE WHEN address_detail != '' AND address_detail IS NOT NULL THEN 1 ELSE 0 END) as total_detail "
        "FROM clients"
    )
    if detail_stats:
        s = detail_stats[0]
        print(f"\n  [상세주소 현황]")
        print(f"    상세주소 있는 건: {s['total_detail']}")
        print(f"    (XX동) 남은 건:  {s['dong_only']}")
        print(f"    (XX동)+상세 건:  {s['dong_plus']}")


if __name__ == "__main__":
    main()
