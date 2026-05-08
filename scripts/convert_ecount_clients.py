#!/usr/bin/env python3
"""
이카운트 ERP 거래처 엑셀 → MES 임포트용 엑셀 변환 스크립트

Usage:
    python3 scripts/convert_ecount_clients.py

Input:  거래처정보.xlsx (프로젝트 루트)
Output: 거래처정보_MES변환.xlsx (프로젝트 루트)
"""

import re
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 설정 ──────────────────────────────────────────────

INPUT_FILE = Path(__file__).parent.parent / "거래처정보.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "거래처정보_MES변환.xlsx"

# 이카운트 헤더 (0-indexed)
# 거래처코드,거래처명,대표자명,업태,종목,Fax,Email,주소1,주소2,(빈열),전화,모바일,검색창내용,계산서발행명,사용구분,이체정보
COL_CODE = 0
COL_NAME = 1
COL_REP = 2
COL_BTYPE = 3
COL_BITEM = 4
COL_FAX = 5
COL_EMAIL = 6
COL_ADDR1 = 7
COL_ADDR2 = 8
# COL_EMPTY = 9  (엑셀에서 추가된 빈 열)
COL_PHONE = 10
COL_MOBILE = 11
COL_SEARCH = 12
COL_INVOICE = 13
COL_ACTIVE = 14
COL_TRANSFER = 15

# ── 계산서발행명 → invoice_type 매핑 ──────────────────

INVOICE_TYPE_MAP = {
    "일별": "PER_ORDER",
    "월별": "MONTHLY",
    "주별": "MONTHLY",
    "15일": "MONTHLY",
    "미분류": "UNDECIDED",
    "카드결재": "CARD",
    "카드결제": "CARD",
    "발행X,카드결재": "CARD",
    "발행X,카드결제": "CARD",
    "타발행": "ISSUED_BY_OTHER",
}

# ── 시/도 목록 (주소 파싱용) ──────────────────────────

SIDO_LIST = [
    "서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시",
    "대전광역시", "울산광역시", "세종특별자치시", "경기도", "강원특별자치도",
    "충청북도", "충청남도", "전북특별자치도", "전라남도", "경상북도", "경상남도",
    "제주특별자치도",
    # 약식
    "충남", "충북", "경북", "경남", "전북", "전남",
    "대전", "대구", "부산", "인천", "광주", "울산", "세종", "서울", "경기", "강원", "제주",
]


def parse_address(addr: str) -> tuple[str, str]:
    """
    주소를 (기본주소, 상세주소)로 분리.

    기본주소: 시도 + 시군구 + 도로명/지번 + (법정동)
    상세주소: 층, 호, 건물명 등

    Returns:
        (base_address, detail_address)
    """
    if not addr or not addr.strip():
        return ("", "")

    addr = addr.strip()
    original = addr

    # 1. 시도 추출
    sido = ""
    for s in SIDO_LIST:
        if addr.startswith(s):
            sido = s
            addr = addr[len(s):].strip()
            break

    # 2. 시군구 추출 (최대 2단계: 천안시 서북구)
    sigungu_parts = []
    for _ in range(2):
        m = re.match(r"^(\S+?[시군구])\s*", addr)
        if m:
            sigungu_parts.append(m.group(1))
            addr = addr[m.end():].strip()

    # 읍면 추출
    m = re.match(r"^(\S+?[읍면])\s*", addr)
    if m:
        sigungu_parts.append(m.group(1))
        addr = addr[m.end():].strip()

    sigungu = " ".join(sigungu_parts)

    # 3. 법정동 (첫 번째 괄호) 추출 — 단, 회사명 괄호 제외
    #    회사명 괄호: (주), (유), (사), (복) 등 1~2글자
    paren_match = None
    beopjeong = ""
    for m in re.finditer(r"\(([^)]+)\)", addr):
        content = m.group(1)
        # 1~2글자 괄호는 회사명이므로 무시
        if len(content) <= 2:
            continue
        paren_match = m
        beopjeong = content
        break

    # 4. 괄호 제거 후 도로명/지번 + 상세 분리
    if paren_match:
        before_paren = addr[: paren_match.start()].strip()
        after_paren = addr[paren_match.end():].strip()
        paren_text = f"({beopjeong})"
    else:
        before_paren = addr
        after_paren = ""
        paren_text = ""

    # 5. 도로명 주소 패턴: ~로/~길/~대로 + 번지
    road_match = re.match(r"^(.+?(?:로|길|대로)\S*\s+\d[\d-]*)", before_paren)
    if road_match:
        road_part = road_match.group(1).strip()
        leftover = before_paren[road_match.end():].strip()
    else:
        # 지번 주소 패턴: 동/리/가 + 번지
        jibun_match = re.match(r"^(\S+?[동리가]\s+\d[\d-]*)", before_paren)
        if jibun_match:
            road_part = jibun_match.group(1).strip()
            leftover = before_paren[jibun_match.end():].strip()
        else:
            # 파싱 실패 — 전체를 기본주소로
            road_part = before_paren
            leftover = ""

    # 6. 기본주소 조립
    base_parts = [sido, sigungu, road_part]
    if paren_text:
        base_parts.append(paren_text)
    base_address = " ".join(p for p in base_parts if p)

    # 7. 상세주소 조립
    detail_parts = [leftover, after_paren]
    detail_address = " ".join(p for p in detail_parts if p).strip()

    # 후처리: 두 번째 괄호가 상세에 남아있을 수 있음 (건물명 등) — 그대로 유지
    return (base_address, detail_address)


def classify_delivery(addr1: str, addr2: str) -> tuple[str, str]:
    """
    주소1과 주소2를 비교하여 배송정보 분류.

    Returns:
        (delivery_method, delivery_address)
    """
    a1 = (addr1 or "").strip()
    a2 = (addr2 or "").strip()

    if not a2 or a1 == a2:
        return ("SAME", "")

    # 화물지점: 짧은 이름 (≤15자)
    if len(a2) <= 15:
        return ("FREIGHT", a2)

    # 별도 주소: 시/도로 시작하는 전체 주소
    prefixes = [
        "서울", "경기", "충청", "전북", "전남", "경상", "강원",
        "대전", "대구", "부산", "인천", "광주", "울산", "세종", "제주",
        "충남", "충북", "경북", "경남", "전라",
    ]
    if any(a2.startswith(p) for p in prefixes):
        return ("DIRECT", a2)

    # 기타 (잡 데이터) — SAME으로 처리
    return ("SAME", "")


def map_invoice_type(value: str) -> str:
    """계산서발행명 → invoice_type 매핑."""
    v = (value or "").strip()
    return INVOICE_TYPE_MAP.get(v, "UNDECIDED")


def safe_str(val) -> str:
    """셀 값을 안전하게 문자열로 변환."""
    if val is None:
        return ""
    return str(val).strip()


def main():
    print(f"입력 파일: {INPUT_FILE}")

    if not INPUT_FILE.exists():
        print(f"오류: {INPUT_FILE} 파일을 찾을 수 없습니다.")
        sys.exit(1)

    print("파일 로딩 중...")
    wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb.active

    # 모든 행을 한번에 읽기 (read_only 모드에서는 iter_rows 사용)
    print("데이터 읽는 중...")
    all_rows = []
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 헤더 찾기
        if header_row_idx is None:
            if row and safe_str(row[0]) == "거래처코드":
                header_row_idx = row_idx
                print(f"헤더 행: {row_idx}")
            continue

        # 헤더 이후 데이터
        code = safe_str(row[0]) if row and len(row) > 0 else ""
        # 사업자번호 패턴이 아니면 건너뜀
        if not re.match(r"^\d{3}-\d{2}-\d{5}$", code):
            continue
        all_rows.append(row)

    wb.close()

    if header_row_idx is None:
        print("오류: '거래처코드' 헤더를 찾을 수 없습니다.")
        sys.exit(1)

    print(f"데이터 행: {len(all_rows)}건 로드 완료")

    # ── 출력 워크북 생성 ──────────────────────────────

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "MES 거래처"

    # 스타일 정의
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # 검증 필요 컬럼 강조색
    verify_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    verify_header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

    # 출력 헤더
    OUT_HEADERS = [
        ("client_code\n(거래처코드=사업자번호)", 20, False),
        ("client_name\n(거래처명)", 25, False),
        ("business_registration_number\n(사업자등록번호)", 22, False),
        ("representative\n(대표자명)", 12, False),
        ("business_type\n(업태)", 15, False),
        ("business_item\n(종목)", 20, False),
        ("phone\n(전화)", 16, False),
        ("mobile\n(모바일)", 16, False),
        ("fax", 16, False),
        ("email", 25, False),
        ("address\n(기본주소) <--검증", 40, True),
        ("address_detail\n(상세주소) <--검증", 20, True),
        ("delivery_method\n(배송방식) <--검증", 14, True),
        ("delivery_address\n(배송지/지점명) <--검증", 20, True),
        ("invoice_type\n(계산서유형) <--검증", 16, True),
        ("invoice_type_원본\n(이카운트 원본값)", 14, False),
        ("search_keywords\n(검색창내용)", 15, False),
        ("transfer_info\n(이체정보)", 12, False),
        ("is_active", 10, False),
        ("addr1_원본\n(주소1 원본)", 40, False),
        ("addr2_원본\n(주소2 원본)", 30, False),
    ]

    # 헤더 쓰기
    for col_idx, (name, width, is_verify) in enumerate(OUT_HEADERS, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = verify_header_fill if is_verify else header_fill
        cell.alignment = header_align
        cell.border = thin_border
        out_ws.column_dimensions[cell.column_letter].width = width

    # 행 높이
    out_ws.row_dimensions[1].height = 35

    # ── 데이터 변환 ──────────────────────────────────

    stats = {
        "total": 0,
        "skipped_inactive": 0,
        "skipped_no_code": 0,
        "written": 0,
        "delivery_same": 0,
        "delivery_freight": 0,
        "delivery_direct": 0,
        "addr_parse_with_detail": 0,
        "addr_parse_no_detail": 0,
    }

    def get_col(row, idx):
        """행에서 컬럼 값을 안전하게 추출."""
        if row and len(row) > idx:
            return safe_str(row[idx])
        return ""

    out_row = 2
    for row in all_rows:
        code = get_col(row, COL_CODE)
        stats["total"] += 1

        # 사용구분 NO 제외
        active = get_col(row, COL_ACTIVE)
        if active.upper() == "NO":
            stats["skipped_inactive"] += 1
            continue

        # 필드 추출
        name = get_col(row, COL_NAME)
        rep = get_col(row, COL_REP)
        btype = get_col(row, COL_BTYPE)
        bitem = get_col(row, COL_BITEM)
        phone = get_col(row, COL_PHONE)
        mobile = get_col(row, COL_MOBILE)
        fax = get_col(row, COL_FAX)
        email = get_col(row, COL_EMAIL)
        addr1 = get_col(row, COL_ADDR1)
        addr2 = get_col(row, COL_ADDR2)
        search = get_col(row, COL_SEARCH)
        invoice_raw = get_col(row, COL_INVOICE)
        transfer_raw = get_col(row, COL_TRANSFER)
        # 이카운트는 "등록" = 이체정보 등록 여부 표시일 뿐, 실제 계좌 아님
        transfer = "" if transfer_raw in ("등록", "") else transfer_raw

        # 주소 파싱
        base_addr, detail_addr = parse_address(addr1)
        if detail_addr:
            stats["addr_parse_with_detail"] += 1
        else:
            stats["addr_parse_no_detail"] += 1

        # 배송정보 분류
        delivery_method, delivery_address = classify_delivery(addr1, addr2)
        stats[f"delivery_{delivery_method.lower()}"] += 1

        # 계산서 유형 매핑
        invoice_type = map_invoice_type(invoice_raw)

        # ── 출력 행 쓰기 ─────────────────────────────

        row_data = [
            code,               # client_code
            name,               # client_name
            code,               # business_registration_number (동일)
            rep,                # representative
            btype,              # business_type
            bitem,              # business_item
            phone,              # phone
            mobile,             # mobile
            fax,                # fax
            email,              # email
            base_addr,          # address (파싱됨)
            detail_addr,        # address_detail (파싱됨)
            delivery_method,    # delivery_method
            delivery_address,   # delivery_address
            invoice_type,       # invoice_type
            invoice_raw,        # invoice_type 원본 (참조용)
            search,             # search_keywords
            transfer,           # transfer_info
            1,                  # is_active
            addr1,              # addr1 원본 (참조용)
            addr2,              # addr2 원본 (참조용)
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = out_ws.cell(row=out_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # 검증 필요 컬럼 배경색
            if OUT_HEADERS[col_idx - 1][2]:
                cell.fill = verify_fill

        out_row += 1
        stats["written"] += 1

    # ── 자동 필터 ─────────────────────────────────────

    out_ws.auto_filter.ref = f"A1:{chr(64 + len(OUT_HEADERS))}{out_row - 1}"

    # ── 틀 고정 (헤더 고정) ───────────────────────────

    out_ws.freeze_panes = "A2"

    # ── 저장 ──────────────────────────────────────────

    out_wb.save(OUTPUT_FILE)
    print(f"\n출력 파일: {OUTPUT_FILE}")
    print(f"\n{'='*50}")
    print(f"  변환 통계")
    print(f"{'='*50}")
    print(f"  전체 거래처:       {stats['total']:>6}건")
    print(f"  사용구분 NO 제외:  {stats['skipped_inactive']:>6}건")
    print(f"  변환 완료:         {stats['written']:>6}건")
    print(f"{'─'*50}")
    print(f"  주소 상세 분리됨:  {stats['addr_parse_with_detail']:>6}건")
    print(f"  주소 상세 없음:    {stats['addr_parse_no_detail']:>6}건")
    print(f"{'─'*50}")
    print(f"  배송 SAME:         {stats['delivery_same']:>6}건")
    print(f"  배송 FREIGHT:      {stats['delivery_freight']:>6}건")
    print(f"  배송 DIRECT:       {stats['delivery_direct']:>6}건")
    print(f"{'='*50}")
    print(f"\n<-- 노란색 컬럼을 눈으로 검증해주세요:")
    print(f"  - address (기본주소): 파싱이 맞는지")
    print(f"  - address_detail (상세주소): 분리가 맞는지")
    print(f"  - delivery_method: FREIGHT/DIRECT 분류가 맞는지")
    print(f"  - invoice_type: 매핑이 맞는지")
    print(f"  - 원본 컬럼(addr1_원본, addr2_원본, invoice_type_원본)과 비교 가능")


if __name__ == "__main__":
    main()
