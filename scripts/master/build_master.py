# -*- coding: utf-8 -*-
# 업로드용 품목 마스터 빌드: 298 base 정규화 + 규격그룹 + 변종전개(예시) + 검토
import sys, os, re, openpyxl
from openpyxl import Workbook
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

SRC = "품목마스터/설계/표준품목_등록구조_수정본.xlsx"
OUT = os.environ.get('MASTER_OUT', "품목마스터/업로드양식/품목업로드_최종.xlsx")

CATEGORIES = {'현수막','배너','깃발·기','시트·스티커','판재출력','출력물','간판','원자재','부속품'}
PRICING = {'AREA','FIXED','GRADE','COMPONENT'}

# 규격그룹 정의
SPEC_GROUPS = {
    '판재두께': ['1T','2T','3T','5T','8T','10T'],
    '호수': ['1호','2호','3호','4호','4-1','5호','6호','7호','7-1','8호','8-1','9호','10호'],
    '원단폭': [],  # TBD — 값 확정 후
}
def value_code(label):
    # 1호->1HO, 4-1->4HO1, 5T->5T
    m = re.match(r'^(\d+)호$', label)
    if m: return m.group(1)+'HO'
    m = re.match(r'^(\d+)-(\d+)$', label)
    if m: return m.group(1)+'HO'+m.group(2)
    return label.replace(' ','')

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["등록데이터_수정본"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else "" for c in rows[0]]
idx = {h: i for i, h in enumerate(hdr)}
def g(r, n):
    i = idx.get(n); return (str(r[i]).strip() if i is not None and i < len(r) and r[i] is not None else "")
data = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
wb.close()

FLAG_BASE = re.compile(r'(원판)$')
FABRIC = re.compile(r'(원단|폰지|사틴|메쉬|친환경|코팅|켄버스|부직포|텐트천)')
# 그룹 과포괄 보정 (2026-06-20 용준님 점검):
#   판재두께 = 포맥스/폼보드/아크릴 계열만 확정. 합판·원목·PC는 표준두께 상이 → 검토.
POMAX_OK = re.compile(r'(포맥스|폼보드|아크릴)')
#   호수 = 깃발·기 GRADE 전부가 아니라, 명백한 비호수(세트·수기대·수기·배너·장식·특호·외건) 제외.
#         나머지는 게양용 표준기 후보 → 용준님 호수 변종 여부 직접 지정.
HO_EXCLUDE = re.compile(r'(수기대|수기$|함세트|지관|특호|배너|만국기|만장기|청사초롱|어구실명제|외\s*건|SET|세트)')

def profile(pm): return pm.replace('*','').strip()
def base_name(name):
    return re.sub(r'\s*(원판|출력)\s*$','',name).strip()

items = []   # 품목마스터(base)
review = []  # 검토필요
for r in data:
    code, name, it, cat = g(r,'item_code'), g(r,'item_name'), g(r,'item_type'), g(r,'대분류')
    pm = profile(g(r,'pricing_method')); unit = g(r,'unit') or 'EA'; gubun = g(r,'구분'); bigo = g(r,'비고/조치')
    igroup = base_name(name)
    # spec_group 마킹 (과포괄 보정 — 명백한 변종만, 의심분은 sg='' + review로 분리)
    sg = ''
    if gubun == '겸업-자재' or FLAG_BASE.search(name):
        if POMAX_OK.search(name):
            sg = '판재두께'
        else:
            review.append(('판재두께 적합성?', code, name, '합판/원목/PC — 표준두께(1·2·3·5·8·10T) 상이 가능, 두께값 확인 후 변종'))
    elif cat == '깃발·기' and g(r,'pricing_method') == 'GRADE*':
        if HO_EXCLUDE.search(name):
            pass  # 호수 아님 → 단일 FIXED (sg='')
        else:
            sg = '호수'
            review.append(('호수 대상 지정', code, name, '게양용 표준기 후보 — 용준님 호수 변종 여부 확정'))
    elif cat == '원자재' and FABRIC.search(name):
        if unit == '롤':
            sg = '원단폭'
            review.append(('원단폭 값 미정', code, name, '폭 mm 목록 확정 후 변종 생성'))
        else:
            review.append(('원단폭 적합성?', code, name, f'unit={unit} — 낱장/개수 판매, 폭 변종 아닐 수 있음'))
    items.append({
        'item_code': code, 'item_name': name, 'item_type': it, 'category': cat,
        'item_group': igroup, 'spec_group': sg, 'pricing_method': pm, 'pricing_profile': pm,
        'unit': unit, 'parent_media': '', 'ecount_code': '', 'base_price': '', 'is_active': 1,
        'variant_base': 1 if sg else 0, 'note': (gubun+(' '+bigo if bigo else '')).strip(),
    })
    if gubun in ('이름확인','간판자재(후속)'):
        review.append((gubun, code, name, bigo))

# 변종전개 예시: 포맥스 원판(두께) — 사용자 명시 케이스만
demo = []
pomax = next((x for x in items if x['item_name']=='포맥스 원판'), None)
if pomax:
    for label in SPEC_GROUPS['판재두께']:
        vc = value_code(label)
        demo.append({'item_code': f"{pomax['item_code']}-{vc}", 'item_name': f"포맥스 {label}",
                     'item_type':'MATERIAL','category':'원자재','item_group':'포맥스',
                     'spec_group':'판재두께','spec_value':label,'pricing_profile':'FIXED','unit':'장'})

# ── Excel 출력 ──
out = Workbook()
s1 = out.active; s1.title = '품목마스터'
cols1 = ['item_code','item_name','item_type','category','item_group','spec_group','pricing_method','pricing_profile','unit','parent_media','ecount_code','base_price','is_active','note']
s1.append(cols1)
for x in items: s1.append([x.get(c,'') for c in cols1])

s2 = out.create_sheet('규격그룹')
s2.append(['spec_group','value_code','label','sort'])
for gname, vals in SPEC_GROUPS.items():
    for i,label in enumerate(vals): s2.append([gname, value_code(label), label, i+1])
    if not vals: s2.append([gname,'(값 미정)','','' ])

s3 = out.create_sheet('변종전개_예시')
cols3 = ['item_code','item_name','item_type','category','item_group','spec_group','spec_value','pricing_profile','unit']
s3.append(cols3)
for x in demo: s3.append([x.get(c,'') for c in cols3])

s4 = out.create_sheet('검토필요')
s4.append(['사유','item_code','item_name','비고'])
for r in review: s4.append(list(r))

out.save(OUT)

# ── 리포트 ──
print(f"생성: {OUT}")
print(f"  품목마스터 {len(items)} / 변종전개예시 {len(demo)} / 검토 {len(review)}")
print(f"  category: {dict(Counter(x['category'] for x in items))}")
print(f"  spec_group 마킹: {dict(Counter(x['spec_group'] for x in items if x['spec_group']))}")
print(f"  변종-base: {sum(x['variant_base'] for x in items)}")
print(f"  검토 사유: {dict(Counter(r[0] for r in review))}")
