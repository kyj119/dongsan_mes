# -*- coding: utf-8 -*-
# 전체 품목 라인 정리 + 주문 수요 — 통합본 609 정본, 대분류×타입×처리방식 + 6개월 주문횟수·매출.
# 용준님 직접 확인용(빠진 품목·불필요 품목·코드 형태). 무형(비품목) 분리.
import sys, os, re, openpyxl
from openpyxl import Workbook
from collections import defaultdict, Counter
sys.stdout.reconfigure(encoding='utf-8')

UNIFIED="품목마스터/설계/동산_품목표준화_통합본.xlsx"
FIXED  ="품목마스터/설계/표준품목_등록구조_수정본.xlsx"
ORDERS =['동산1.1~3.3.xlsx','동산3.3~4.15.xlsx','동산4.15~6.12.xlsx','선명(2)1.1~6.13.xlsx']
OUT    =os.environ.get('ORG_OUT', "품목마스터/업로드양식/품목정리_전체.xlsx")

CAT_MAP={'UV출력':'출력물','솔벤출력':'출력물','부직포출력':'출력물','평판출력':'판재출력','평판':'판재출력',
 '시트(수성/솔벤)':'시트·스티커','전사자재':'깃발·기','태극기':'깃발·기','깃발':'깃발·기','새마을기':'깃발·기',
 '민방위기':'깃발·기','태극기원단':'원자재','전사원단':'원자재','실사원단(수성/솔벤)':'원자재','원재료':'원자재',
 '배너대':'배너','윈드배너':'배너','가로등배너':'배너','에어간판':'간판','간판':'간판','(선명:간판)':'간판',
 '간판자재':'원자재','잉크':'원자재','부직포':'원자재','부자재':'부속품','장비판매':'(설비)','미분류':'(검토)'}
INTANGIBLE_GRP={'기타, 할인','용차,퀵','설치비','(선명:영업부)'}
AREA_GRP={'UV출력','솔벤출력','부직포출력','시트(수성/솔벤)'}

def norm(n):
    n=re.sub(r'\s*\[[^\]]*\]','',str(n)); n=re.sub(r'\s*<[^>]*>','',n)
    return re.sub(r'\s+',' ',n).strip()
def num(v):
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).replace(',','').strip())
    except: return 0.0
def load(path, sheet):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws=wb[sheet]; rows=list(ws.iter_rows(values_only=True))
    H=[str(c).strip() if c else '' for c in rows[0]]; I={h:i for i,h in enumerate(H)}
    d=[r for r in rows[1:] if any(c is not None for c in r)]; wb.close(); return I,d

# 주문 수요 집계
o_cnt=Counter(); o_rev=Counter()
for f in ORDERS:
    wb=openpyxl.load_workbook('품목마스터/원천주문데이터/'+f, read_only=True, data_only=True)
    ws=wb['주문서현황내역']
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i<2 or row[3] is None: continue
        nm=norm(row[3])
        if nm: o_cnt[nm]+=1; o_rev[nm]+=num(row[8])
    wb.close()

# 수정본 298 → 코드/타입/대분류
I2,d2=load(FIXED,'등록데이터_수정본')
def g2(r,n):
    i=I2.get(n); return (str(r[i]).strip() if i is not None and i<len(r) and r[i] is not None else '')
fixed_by_name={}
for r in d2: fixed_by_name.setdefault(g2(r,'item_name'),(g2(r,'item_code'),g2(r,'item_type'),g2(r,'대분류')))

# 통합본 정본 + 변종예시 매핑
I1,d1=load(UNIFIED,'1.표준품목마스터')
def g1(r,n):
    i=I1.get(n); return (str(r[i]).strip() if i is not None and i<len(r) and r[i] is not None else '')
ex2std={}
for r in d1:
    std=g1(r,'표준 품목명')
    for e in g1(r,'변종 예시').split(','):
        e=norm(e)
        if e: ex2std.setdefault(e,std)
    ex2std.setdefault(norm(std),std)
# 주문 고유명 → 표준품목 귀속
std_names=sorted(set(g1(r,'표준 품목명') for r in d1), key=len, reverse=True)
std_cnt=Counter(); std_rev=Counter()
for nm,c in o_cnt.items():
    s=ex2std.get(nm)
    if not s:
        for cand in std_names:
            if len(cand)>=2 and (nm==cand or nm.startswith(cand+' ') or nm.startswith(cand+'(')): s=cand; break
    if s: std_cnt[s]+=c; std_rev[s]+=o_rev[nm]

PFX={'PRODUCT':'P','GOODS':'G','MATERIAL':'M'}; seq=defaultdict(int)
items=[]; intangible=[]
for r in d1:
    grp=g1(r,'계층그룹(분류)'); nm=g1(r,'표준 품목명'); role=g1(r,'역할'); jp=g1(r,'자재/제품')
    try: vc=int(float(g1(r,'변종수') or 0))
    except: vc=0
    src=g1(r,'출처')
    if grp in INTANGIBLE_GRP: intangible.append((grp,nm,vc,src,std_cnt[nm],int(std_rev[nm]))); continue
    cat=CAT_MAP.get(grp,'(검토)')
    if nm in fixed_by_name:
        code,itype,fcat=fixed_by_name[nm]; cat=fcat or cat; status='확정(298)'
    else:
        itype='MATERIAL' if ('자재' in jp or '자재' in role) else 'PRODUCT'
        seq[itype]+=1; code=f"{PFX[itype]}-(신규{seq[itype]:03})"; status='신규(추가대상)'
    if grp in AREA_GRP: how='면적단가→소재흡수'
    elif cat=='원자재' and grp=='잉크': how='변종(색상)?'
    elif grp=='간판자재': how='자재(두께변종?)'
    elif itype=='MATERIAL': how='자재단일'
    elif vc>1: how='변종전개'
    else: how='단일'
    cnt=std_cnt[nm]; rev=int(std_rev[nm])
    # 수요판정 (자재는 매입용이라 빈도 무관)
    if itype=='MATERIAL': demand='자재(매입·빈도무관)'
    elif cnt==0: demand='0회·죽은품목후보(계절확인)'
    elif cnt<=2: demand='1~2회·ad-hoc후보'
    else: demand='활성'
    items.append({'대분류':cat,'타입':itype,'코드':code,'품목명':nm,'계층그룹':grp,'변종수':vc,
                  '처리방식':how,'주문6M':cnt,'매출6M':rev,'수요판정':demand,'출처':src,'상태':status})

out=Workbook(); s=out.active; s.title='전체품목'
cols=['대분류','타입','코드','품목명','계층그룹','변종수','처리방식','주문6M','매출6M','수요판정','출처','상태']
s.append(cols)
order={'현수막':1,'배너':2,'깃발·기':3,'시트·스티커':4,'판재출력':5,'출력물':6,'간판':7,'원자재':8,'부속품':9,'(설비)':10,'(검토)':11}
for x in sorted(items, key=lambda x:(order.get(x['대분류'],99), x['타입'], -x['매출6M'])):
    s.append([x[c] for c in cols])
s2=out.create_sheet('무형(비품목)')
s2.append(['계층그룹','품목명','변종수','출처','주문6M','매출6M','처리'])
for grp,nm,vc,src,c,rv in intangible: s2.append([grp,nm,vc,src,c,rv,'주문라인 자유입력'])
out.save(OUT)

# 콘솔 요약
print(f'생성: {OUT}  (품목 {len(items)} / 무형 {len(intangible)})')
sell=[x for x in items if x['타입']!='MATERIAL']
dead=[x for x in sell if x['주문6M']==0]; low=[x for x in sell if 1<=x['주문6M']<=2]
ranked=sorted(sell, key=lambda x:-x['매출6M']); srev=sum(x['매출6M'] for x in sell) or 1
print(f'판매품 {len(sell)}: 0회 {len(dead)} / 1~2회 {len(low)} / 활성 {len(sell)-len(dead)-len(low)}')
pareto=' / '.join('상위%d=%.0f%%'%(k,100*sum(x['매출6M'] for x in ranked[:k])/srev) for k in (20,50,100))
print('매출 파레토:', pareto)
print('0회 죽은품목 계층그룹:', dict(Counter(x['계층그룹'] for x in dead).most_common(8)))
print('처리방식:', dict(Counter(x['처리방식'] for x in items)))
