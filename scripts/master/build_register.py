# -*- coding: utf-8 -*-
# 활성 제품 등록 — 수정본 298 정본(대분류·코드 확정) 중 PRODUCT & 주문 1회+. 마이그 0327 + 미리보기.
# 기준(2026-06-20 용준님): 죽은품목(0회)·매입전용 자재·상품 = 보류(별도 검토). 명백 판매제품만 우선.
# 0325(마스터 로드)와 동일 코드·pricing. 수요는 주문 18,491건에서 item_name 매칭.
import openpyxl, sys, os, re
from openpyxl import Workbook
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

FIXED="품목마스터/설계/표준품목_등록구조_수정본.xlsx"
ORDERS=['동산1.1~3.3.xlsx','동산3.3~4.15.xlsx','동산4.15~6.12.xlsx','선명(2)1.1~6.13.xlsx']
MIG="migrations/0327_item_register_active.sql"
PREVIEW=os.environ.get('REG_OUT',"품목마스터/업로드양식/등록대상_미리보기.xlsx")

def norm(n):
    n=re.sub(r'\s*\[[^\]]*\]','',str(n)); n=re.sub(r'\s*<[^>]*>','',n); return re.sub(r'\s+',' ',n).strip()
def sqlstr(s): return "'"+str(s).replace("'","''")+"'"
def profile(pm):  # 수정본 pricing_method → (method CHECK FIXED/AREA, profile)
    pm=pm.replace('*','').strip()
    if pm=='AREA': return 'AREA','AREA'
    if pm=='GRADE': return 'FIXED','GRADE'
    if pm=='COMPONENT': return 'FIXED','COMPONENT'
    return 'FIXED','FIXED'

# 주문 수요
o_cnt=Counter()
for f in ORDERS:
    wb=openpyxl.load_workbook('품목마스터/원천주문데이터/'+f, read_only=True, data_only=True)
    for i,row in enumerate(wb['주문서현황내역'].iter_rows(values_only=True)):
        if i<2 or row[3] is None: continue
        nm=norm(row[3])
        if nm: o_cnt[nm]+=1
    wb.close()

# 수정본 298
wb=openpyxl.load_workbook(FIXED, read_only=True, data_only=True)
ws=wb['등록데이터_수정본']; rows=list(ws.iter_rows(values_only=True))
H=[str(c).strip() if c else '' for c in rows[0]]; I={h:i for i,h in enumerate(H)}
def g(r,n):
    i=I.get(n); return (str(r[i]).strip() if i is not None and i<len(r) and r[i] is not None else '')
data=[r for r in rows[1:] if any(c is not None for c in r)]; wb.close()

# item_name → 주문 수요: 정확 + substring(긴이름) + 경계 접미사(-S/+돔보/규격 등)
BND=set(' ([-/+,)')
def demand(name):
    tot=0
    for on,oc in o_cnt.items():
        if on==name or (len(name)>=4 and name in on) or \
           (on.startswith(name) and len(on)>len(name) and on[len(name)] in BND):
            tot+=oc
    return tot

reg=[]; hold=[]
for r in data:
    code=g(r,'item_code'); name=g(r,'item_name'); itype=g(r,'item_type'); cat=g(r,'대분류')
    pm,pp=profile(g(r,'pricing_method')); unit=g(r,'unit') or 'EA'; gubun=g(r,'구분')
    cnt=demand(name)
    if itype=='PRODUCT' and cnt>=1 and gubun not in ('이름확인','간판자재(후속)'):
        reg.append({'code':code,'name':name,'cat':cat,'pm':pm,'pp':pp,'unit':unit,'cnt':cnt})
    else:
        why=('0회죽은품목' if itype=='PRODUCT' and cnt==0 else
             ('이름확인' if gubun=='이름확인' else
              ('간판자재후속' if gubun=='간판자재(후속)' else
               (f'{itype}(보류)' if itype!='PRODUCT' else '보류'))))
        hold.append({'code':code,'name':name,'cat':cat,'itype':itype,'why':why,'cnt':cnt})

# 마이그
lines=["-- 0327: 활성 판매제품 1차 등록 (수정본 298 중 PRODUCT & 주문 6개월 1회+)",
 "-- 기준: 죽은품목(0회)·자재·상품·이름확인·간판자재 = 보류(별도 검토). 명백 판매제품만.",
 "-- 0325와 동일 코드·pricing. is_active=0 staged(라이브 미노출). category_id=NAME 서브쿼리.",
 "-- 비파괴: item_code NOT EXISTS. 변종(두께/호수)·소재흡수는 후속.",""]
for x in reg:
    lines.append(
    f"INSERT INTO items (item_code,item_name,item_type,category_id,pricing_method,pricing_profile,is_sales_item,is_purchase_item,production_required,unit,is_active)\n"
    f"SELECT {sqlstr(x['code'])},{sqlstr(x['name'])},'PRODUCT',(SELECT id FROM item_categories WHERE category_name={sqlstr(x['cat'])}),{sqlstr(x['pm'])},{sqlstr(x['pp'])},1,0,1,{sqlstr(x['unit'])},0\n"
    f"WHERE NOT EXISTS (SELECT 1 FROM items WHERE item_code={sqlstr(x['code'])});")
with open(MIG,'w',encoding='utf-8') as f: f.write("\n".join(lines)+"\n")

# 미리보기
out=Workbook(); s=out.active; s.title='등록대상'
s.append(['코드','대분류','품목명','단가방식','주문6M','단위'])
for x in sorted(reg,key=lambda x:(x['cat'],-x['cnt'])): s.append([x['code'],x['cat'],x['name'],f"{x['pm']}/{x['pp']}",x['cnt'],x['unit']])
s2=out.create_sheet('보류')
s2.append(['코드','대분류','품목명','타입','사유','주문6M'])
for x in sorted(hold,key=lambda x:(x['why'],-x['cnt'])): s2.append([x['code'],x['cat'],x['name'],x['itype'],x['why'],x['cnt']])
out.save(PREVIEW)

print(f"마이그: {MIG} ({len(reg)} INSERT)")
print(f"미리보기: {PREVIEW}")
print(f"등록 {len(reg)} / 보류 {len(hold)}")
print("등록 대분류:", dict(Counter(x['cat'] for x in reg)))
print("보류 사유:", dict(Counter(x['why'] for x in hold)))
