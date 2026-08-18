# -*- coding: utf-8 -*-
"""
재고 실사표 생성기 (법인별 xlsx) — 부분실사(ABC) 전용

왜 필요한가 — MES 재고는 못 쓴다. `inventory_transactions` 가 **전체 1건**이고 자동차감 발동도 0건이라
  수량이 있어도 특정 시점 재고가 아니다(2026-08-18 실측). 그래서 손익의 매출원가가 재고 증감을 반영하지
  못하고, 신설 사업장일수록 왜곡이 크다(청주 실측: 반영 전 -5,654만 → 후 -1,692만).
  세무장부 재고자산도 실물과 안 맞는다(용준님 확인). ⇒ **실물을 세는 것 말고 방법이 없다.**

전수는 불가능하지만 전수일 필요도 없다 — 실측 ABC:
  · 동산 431품목 중 **56품목이 매입액의 80%** · 선명 359품목 중 76품목
  · A(80%) + B(95%) 까지만 세면 금액의 대부분이 잡힌다. C 는 세지 않는다.

★위치 열을 「기입」으로 비워 두는 게 설계 핵심이다.
  지금 `inventory_locations` 가 0건이라 어느 자재가 어느 방에 있는지 시스템이 모른다.
  첫 회차에 세면서 적게 하면 그게 곧 매핑이 되고, 다음 회차부터 **위치순으로 정렬된 표**를 낼 수 있다.
  즉 실사가 매핑을 만든다. 매핑을 먼저 만들려고 하면 시작을 못 한다.

사용:
  python scripts/stock-count-sheet.py                      # e1·e2 · 누적 95% 까지
  python scripts/stock-count-sheet.py --entity 1
  python scripts/stock-count-sheet.py --pct 80             # A 등급만(가장 짧은 실사)
  python scripts/stock-count-sheet.py --months 6           # 최근 6개월 매입 기준으로 순위
  python scripts/stock-count-sheet.py --out "Z:/Designs"

★ 읽기 전용이다. prod 에 아무것도 쓰지 않는다.
"""
import sys, io, os, json, subprocess, argparse, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import warnings; warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENTITY = {1: '동산기획', 2: '선명커뮤니케이션', 3: '동산기획(청주)'}

ap = argparse.ArgumentParser()
ap.add_argument('--entity', type=int, default=0, help='1|2|3 (미지정=1,2 둘 다)')
ap.add_argument('--pct', type=float, default=95, help='누적 매입액 몇 %%까지 표에 넣을지')
ap.add_argument('--months', type=int, default=12, help='순위 산정에 쓸 매입 기간(개월)')
ap.add_argument('--out', default='docs/analysis', help='출력 폴더')
ap.add_argument('--date', default='', help='실사 기준일(미지정=지난달 말일)')
A = ap.parse_args()

# ── 기준일: 미지정이면 지난달 말일(실사는 통상 월말 마감 기준)
if A.date:
    BASE = A.date
else:
    t = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    BASE = t.isoformat()
FROM = (datetime.date.fromisoformat(BASE) - datetime.timedelta(days=A.months * 31)).isoformat()


def d1(sql):
    """prod D1 SELECT — wrangler 경유(다른 진단 스크립트와 같은 축).
    ★shell=True 라 개행이 있으면 `--command` 가 끊긴다. 반드시 한 줄로 접고 큰따옴표로 감쌀 것."""
    one = ' '.join(sql.split())
    cmd = ('npx wrangler d1 execute webapp-production --remote --json --command '
           '"' + one.replace('"', '\\"') + '"')
    out = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', shell=True).stdout
    i = out.find('[\n  {')
    if i < 0:
        raise SystemExit('wrangler 응답 해석 실패:\n' + out[:400])
    return json.loads(out[i:])[0]['results']


def fetch(e):
    return d1(f"""
      SELECT i.item_code cd, i.item_name nm, COALESCE(i.specification,'') spec,
             COALESCE(i.unit,'') un, COALESCE(i.category,'(미분류)') cat,
             COALESCE(i.avg_unit_cost,0) cost,
             CAST(SUM(oi.amount) AS INT) amt, COUNT(*) n
        FROM purchase_orders po
        JOIN purchase_order_items oi ON oi.po_id = po.id
        JOIN items i ON i.id = oi.item_id
       WHERE po.entity_id = {e} AND po.status NOT IN ('DRAFT','CANCELLED')
         AND po.order_date >= '{FROM}' AND po.order_date <= '{BASE}'
         AND i.is_active = 1
       GROUP BY i.id ORDER BY amt DESC""")


def zones(e):
    rows = d1(f"SELECT zone_name nm, COALESCE(zone_code,'') cd FROM storage_zones "
              f"WHERE entity_id = {e} AND COALESCE(is_active,1) = 1 ORDER BY sort_order, id")
    return [f"{r['nm']}({r['cd']})" if r['cd'] else r['nm'] for r in rows]


# ── 스타일
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FILL = PatternFill('solid', fgColor='1F3864')
H_FONT = Font(bold=True, color='FFFFFF', size=10)
IN_FILL = PatternFill('solid', fgColor='FFF2CC')   # 사람이 적는 칸
WARN_FILL = PatternFill('solid', fgColor='FCE4E4')  # 단가 없음
GRADE_FILL = {'A': PatternFill('solid', fgColor='C6E0B4'), 'B': PatternFill('solid', fgColor='FFE699')}

# ★열 구성의 핵심 = 「① 롤·포장 수」와 「② 롤당 수량」을 분리한 것.
#   원단 단가는 **야드당**인데 사람은 **롤 단위**로 센다(청주 실사표도 「수량(롤)」·「야드」 2열이었다).
#   ①만 적고 끝내면 야드가 아니라 롤 수가 금액에 곱해져 수백 배 틀린다.
#   길이 단위(yd·m·㎡)만 ②를 비워 두고, 개수 단위(EA·롤·통·장)는 ②=1 을 미리 채운다.
COLS = [('No', 5), ('등급', 6), ('품목코드', 15), ('품목명', 32), ('규격', 15), ('단위', 6),
        ('보관위치', 13), ('① 롤·포장 수', 12), ('② 롤당 수량', 12), ('실사수량 ①×②', 13),
        ('단가', 11), ('금액', 14), ('비고', 14)]
LENGTH_UNITS = {'yd', 'm', '㎡', 'M', 'YD'}


def build(e, rows, out_dir):
    total = sum(r['amt'] for r in rows) or 1
    cum, graded = 0, []
    for r in rows:
        cum += r['amt']
        p = cum / total * 100
        g = 'A' if p <= 80 else ('B' if p <= 95 else 'C')
        if p - (r['amt'] / total * 100) >= A.pct:      # 컷 초과분은 버린다
            break
        graded.append((g, r))

    by_cat = {}
    for g, r in graded:
        by_cat.setdefault(r['cat'], []).append((g, r))
    # 분류 정렬 = 그 분류의 매입액 합 큰 순
    order = sorted(by_cat, key=lambda c: -sum(r['amt'] for _, r in by_cat[c]))
    zlist = zones(e)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 분류별 시트
    for cat in order:
        ws = wb.create_sheet(cat[:28])
        ws.sheet_view.showGridLines = False
        ws['A1'] = f'재고 실사표 — {ENTITY[e]} · {cat}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'실사 기준일  {BASE}       실사자 ____________       작성일 ____________'
        ws['A2'].font = Font(size=10, color='555555')
        ws['A3'] = ('※ 노란 칸만 적으면 실사수량·금액은 자동 계산됩니다. '
                    '★원단은 ①에 롤 수, ②에 롤당 야드를 적으세요(단가가 야드당이라 ②를 빼면 금액이 틀립니다). '
                    '개수 품목은 ②가 1로 채워져 있으니 ①만 적으면 됩니다.')
        ws['A3'].font = Font(size=9, color='C00000')
        if zlist:
            ws['A4'] = '보관위치 예: ' + ' · '.join(zlist) + '  (없으면 자유 기입)'
            ws['A4'].font = Font(size=9, color='555555')

        hr = 6
        for c, (name, w) in enumerate(COLS, start=1):
            cell = ws.cell(hr, c, name)
            cell.fill, cell.font, cell.border = H_FILL, H_FONT, BOX
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[hr].height = 22

        r0 = hr + 1
        for idx, (g, it) in enumerate(by_cat[cat], start=1):
            r = r0 + idx - 1
            cost = float(it['cost'] or 0)
            per = None if it['un'] in LENGTH_UNITS else 1     # 개수 단위는 ②=1 프리필
            vals = [idx, g, it['cd'], it['nm'], it['spec'], it['un'], '', None, per, None,
                    cost if cost else None, None, '' if cost else '단가확인']
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.border = BOX
                cell.alignment = Alignment(horizontal='center' if c in (1, 2, 6, 7) else 'left',
                                           vertical='center', wrap_text=(c == 4))
            ws.cell(r, 2).fill = GRADE_FILL.get(g, PatternFill())
            for c in (7, 8, 9):
                ws.cell(r, c).fill = IN_FILL
            ws.cell(r, 10).value = f'=IF(H{r}="","",H{r}*IF(I{r}="",1,I{r}))'
            ws.cell(r, 12).value = f'=IF(J{r}="","",J{r}*K{r})'
            for c in (8, 9, 10, 11, 12):
                ws.cell(r, c).number_format = '#,##0.##'
                ws.cell(r, c).alignment = Alignment(horizontal='right', vertical='center')
            if not cost:
                ws.cell(r, 11).fill = WARN_FILL
                ws.cell(r, 13).fill = WARN_FILL
            ws.row_dimensions[r].height = 19

        rl = r0 + len(by_cat[cat])
        ws.cell(rl, 4, '소　계').font = Font(bold=True)
        ws.cell(rl, 12, f'=SUM(L{r0}:L{rl - 1})')
        for c in range(1, len(COLS) + 1):
            ws.cell(rl, c).border = BOX
            ws.cell(rl, c).fill = PatternFill('solid', fgColor='D9E1F2')
        ws.cell(rl, 12).number_format = '#,##0'
        ws.cell(rl, 12).font = Font(bold=True)
        ws.cell(rl, 12).alignment = Alignment(horizontal='right')

        # 종이로 들고 다니며 센다 — 인쇄 설정이 실제로 중요하다
        ws.print_title_rows = f'{hr}:{hr}'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.freeze_panes = ws.cell(r0, 1)

    # ── 요약 시트(맨 앞)
    sm = wb.create_sheet('요약', 0)
    sm.sheet_view.showGridLines = False
    sm['A1'] = f'재고 실사 요약 — {ENTITY[e]}'
    sm['A1'].font = Font(bold=True, size=15)
    sm['A2'] = (f'기준일 {BASE} · 순위산정 매입기간 {FROM} ~ {BASE}({A.months}개월) · '
                f'누적 {A.pct:g}% 컷 · 대상 {len(graded)}품목 / 전체 {len(rows)}품목')
    sm['A2'].font = Font(size=10, color='555555')
    for c, (name, w) in enumerate([('분류', 22), ('품목수', 10), ('A등급', 9), ('B등급', 9), ('실사금액', 16)], 1):
        cell = sm.cell(4, c, name)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, BOX
        cell.alignment = Alignment(horizontal='center')
        sm.column_dimensions[get_column_letter(c)].width = w
    for i2, cat in enumerate(order, start=5):
        items = by_cat[cat]
        rl = 6 + len(items)                      # 각 시트의 소계행
        sm.cell(i2, 1, cat).border = BOX
        sm.cell(i2, 2, len(items)).border = BOX
        sm.cell(i2, 3, sum(1 for g, _ in items if g == 'A')).border = BOX
        sm.cell(i2, 4, sum(1 for g, _ in items if g == 'B')).border = BOX
        cell = sm.cell(i2, 5, f"='{cat[:28]}'!L{rl}")
        cell.border, cell.number_format = BOX, '#,##0'
    last = 4 + len(order)
    sm.cell(last + 1, 1, '합　계').font = Font(bold=True)
    cell = sm.cell(last + 1, 5, f'=SUM(E5:E{last})')
    cell.font, cell.number_format = Font(bold=True), '#,##0'
    for c in range(1, 6):
        sm.cell(last + 1, c).fill = PatternFill('solid', fgColor='D9E1F2')
        sm.cell(last + 1, c).border = BOX
    sm.cell(last + 3, 1, '읽는 법').font = Font(bold=True)
    for i3, t in enumerate([
        '· A등급 = 매입액 누적 80%까지. 시간이 없으면 A만 세도 금액의 8할이 잡힙니다.',
        '· B등급 = 80~95%. C등급(나머지 5%)은 이 표에 없습니다 — 세지 않습니다.',
        '· 「보관위치」를 적어 주시면 다음 회차부터 위치순으로 정렬된 표를 드립니다.',
        '· 단가는 최근 매입 평균(야드당·개당)입니다. 빨간 칸은 단가가 없어 금액이 안 나옵니다.',
        '· ★원단은 ①롤 수 + ②롤당 야드를 함께 적으세요. ②를 빼면 금액이 수백 배 작게 나옵니다.',
        '· 세지 못한 품목은 수량을 비워 두세요. 0 을 적으면 「재고 없음」으로 확정됩니다.',
    ]):
        sm.cell(last + 4 + i3, 1, t).font = Font(size=10)

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f'재고실사표_{ENTITY[e]}_{BASE.replace("-", "")}.xlsx')
    wb.save(path)
    cover = sum(r['amt'] for _, r in graded) / total * 100
    print(f'  {ENTITY[e]:<18} {len(graded):>3}품목 (A {sum(1 for g, _ in graded if g == "A")}) '
          f'· 매입액 {cover:.1f}% 커버 · {len(order)}개 분류 → {path}')
    return path


targets = [A.entity] if A.entity else [1, 2]
print(f'재고 실사표 생성 · 기준일 {BASE} · 매입 {FROM}~{BASE}\n')
for e in targets:
    rows = fetch(e)
    if not rows:
        print(f'  {ENTITY[e]}: 매입 실적이 없어 표를 만들 수 없습니다.'); continue
    build(e, rows, A.out)
print('\n※ 「보관위치」 열은 비워 두었습니다 — 첫 회차에 적으신 값이 다음 회차 정렬 기준이 됩니다.')
