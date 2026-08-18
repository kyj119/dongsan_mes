# -*- coding: utf-8 -*-
"""이카운트 「거래처등록」 엑셀 → MES clients 미등록분 추가.

무엇을 하나
  엑셀(거래처코드·거래처명·대표자명·전화·모바일·검색창내용·사용구분·이체정보)을 읽어
  MES `clients` 에 **없는 것만** INSERT SQL 로 낸다. 기본은 dry-run.

왜 이렇게 매칭하나
  1순위 = **사업자번호**(client_code / business_registration_number). 이름은 표기 변종이 많다.
  2순위 = 정규화 이름(법인 접두어·공백·구두점 제거) — importer 의 nrm() 과 **같은 규칙**이어야
          "거래처를 등록했는데 주문 importer 는 여전히 못 찾는" 상태가 안 생긴다.

⚠️ 개인 거래처의 거래처코드는 **주민등록번호**일 수 있다(13자리). 그건 저장하지 않는다 —
   대체 코드(C-{일련})를 부여하고 사업자번호 칸은 비운다. 원본 값은 화면·로그·파일 어디에도 남기지 않는다.

⚠️ 사용구분 != YES 는 제외한다(미사용 거래처). --include-unused 로만 뒤집을 수 있다.

사용법
  python scripts/client-master-sync.py                       # 대조만 (SQL 미생성)
  python scripts/client-master-sync.py --out new_clients.sql # SQL 파일 생성
"""
import argparse
import io
import json
import os
import re
import subprocess
import sys
from collections import Counter

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_SRC = os.path.join(ROOT, '품목마스터', '매핑', '동산기획 거래처목록_26.08.13.xlsx')

RX_LEGAL = re.compile(r'\(주\)|주식회사|㈜|유한회사|주\)')
RX_BIZNO = re.compile(r'^\d{3}-\d{2}-\d{5}$')      # 사업자번호 (저장 가능)
RX_ECOUNT = re.compile(r'^\d{5}$')                 # 이카운트 내부코드 — MES 기존 관례(00013 …)
RX_RRN = re.compile(r'^\d{6}-\d{7}$')              # 주민등록번호 (저장 금지)


RX_BIZ_RAW = re.compile(r'^\d{10}$')               # 대시 없는 사업자번호 (원본 입력 흔들림)


def fix_row(r):
    """원본 입력 오류를 교정한다. 반환 (종류, client_code, 사업자번호, 거래처명).

    이카운트 원본에 실제로 섞여 있던 것들 —
      · 거래처코드 칸에 상호, 거래처명 칸에 사업자번호  → **뒤바뀜**. 그대로 넣으면 이름이 숫자가 된다.
      · 대시 없는 10자리(3080770452)                → 사업자번호로 정규화
      · 날짜(2024-05-24)·사업자번호+접미(...-1)      → 사업자번호로 볼 수 없다 → 대체코드
    ⚠️ 주민번호형은 **어느 칸에도 넣지 않는다**(대체코드는 호출부가 채운다).
    """
    code, name = r['code'], r['name']
    if not RX_BIZNO.match(code) and RX_BIZNO.match(name):
        code, name = name, code          # 코드↔명 반전 교정
    if RX_BIZ_RAW.match(code):           # 3080770452 → 308-07-70452
        code = f'{code[:3]}-{code[3:5]}-{code[5:]}'

    if RX_BIZNO.match(code):
        return 'BIZNO', code, code, name
    if RX_ECOUNT.match(code):
        return 'ECOUNT', code, None, name   # MES가 이미 쓰는 형식 그대로 — 새 관례를 만들지 않는다
    if RX_RRN.match(code):
        return 'RRN', None, None, name      # 저장 금지
    if code:
        return 'OTHER', None, None, name    # 날짜·접미 등 의미 없는 값 → 대체코드로 대체
    return 'NONE', None, None, name


def nrm(s):
    """비교용 정규화 — scripts/ecount-order-import.py 의 nrm() 과 반드시 동일해야 한다."""
    s = RX_LEGAL.sub('', str(s or '').strip().lower())
    return re.sub(r'[\s\-_.,/()\[\]]+', '', s)


def q(v):
    """SQL 리터럴. None/'' → NULL. ⚠️ 'null' 문자열이 되지 않도록 None 을 먼저 거른다."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"


def d1(sql):
    """prod D1 읽기. --json 앞에 붙는 배너를 잘라낸다."""
    out = subprocess.run(
        ['npx', 'wrangler', 'd1', 'execute', 'webapp-production', '--remote', '--json', '--command', sql],
        capture_output=True, text=True, encoding='utf-8', shell=True)
    if '[' not in out.stdout:
        print(out.stdout or out.stderr, file=sys.stderr)
        raise SystemExit('D1 조회 실패')
    return json.loads(out.stdout[out.stdout.index('['):])[0]['results']


def load_excel(path, include_unused):
    """거래처등록 시트 → 행 리스트. read_only 는 dimension 이 깨져 있어 쓰지 않는다."""
    ws = openpyxl.load_workbook(path, data_only=True)['거래처등록']
    hdr = {}
    for j in range(1, ws.max_column + 1):
        name = str(ws.cell(2, j).value or '').strip()
        if name:
            hdr[name] = j
    need = ['거래처코드', '거래처명', '사용구분']
    missing = [k for k in need if k not in hdr]
    if missing:
        raise SystemExit(f'엑셀 헤더 누락: {missing} (2행이 헤더여야 함)')

    def cell(i, name):
        j = hdr.get(name)
        return str(ws.cell(i, j).value or '').strip() if j else ''

    rows, unused = [], 0
    for i in range(3, ws.max_row + 1):
        name = cell(i, '거래처명')
        if not name:
            continue
        if not include_unused and cell(i, '사용구분').upper() != 'YES':
            unused += 1
            continue
        rows.append({
            'code': cell(i, '거래처코드'),
            'name': name,
            'rep': cell(i, '대표자명'),
            'phone': cell(i, '전화'),
            'mobile': cell(i, '모바일'),
            'keywords': cell(i, '검색창내용'),
            'transfer': cell(i, '이체정보'),
        })
    return rows, unused


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--src', default=DEFAULT_SRC)
    ap.add_argument('--out', help='INSERT SQL 출력 경로 (미지정 시 대조만)')
    ap.add_argument('--include-unused', action='store_true', help='사용구분 NO 도 포함(기본 제외)')
    a = ap.parse_args()

    rows, unused = load_excel(a.src, a.include_unused)
    # ★교정을 **매칭 전에** 끝낸다 — 뒤바뀐 행을 원본대로 대조하면 기존 거래처를 못 찾고
    #   중복 INSERT 가 된다(코드도 이름도 엉뚱한 값으로 조회하게 되므로).
    fixed = 0
    for r in rows:
        r['kind'], r['code2'], r['bizno'], name2 = fix_row(r)
        if name2 != r['name'] or (r['code2'] and r['code2'] != r['code']):
            fixed += 1
        r['name'] = name2
    print(f'[엑셀] 사용 {len(rows)}곳 · 미사용 제외 {unused}곳 · 원본 입력오류 교정 {fixed}곳')

    # ── MES 현황 ──────────────────────────────────────────────────────────
    mes = d1('SELECT id, client_code, client_name, business_registration_number, search_keywords, '
             'COALESCE(is_active,1) AS act FROM clients')
    by_code, by_name, alias_of = {}, {}, {}
    for r in mes:
        for c in (r.get('client_code'), r.get('business_registration_number')):
            if c and str(c).strip():
                by_code[str(c).strip()] = r
        by_name.setdefault(nrm(r['client_name']), r)
        # 별칭도 이름 인덱스에 넣는다 — importer(cidx)가 search_keywords 를 쓰기 때문에
        # 여기서 빼면 "이미 별칭으로 잡히는 곳"을 신규로 잘못 세게 된다.
        kws = [k.strip() for k in str(r.get('search_keywords') or '').replace(',', '|').split('|') if k.strip()]
        alias_of[r['id']] = {nrm(k) for k in kws}
        for k in kws:
            by_name.setdefault(nrm(k), r)
    print(f'[MES ] 거래처 {len(mes)}곳 (코드 인덱스 {len(by_code)} · 이름·별칭 인덱스 {len(by_name)})')

    # ── 차집합 ────────────────────────────────────────────────────────────
    # 별칭 후보 = 사업자번호는 같은데 **이카운트 상호가 MES 이름·별칭 어디에도 없는** 경우.
    #   importer 는 이름으로만 찾으므로(cidx) 이걸 안 채우면 그 거래처 주문이 통째로 스킵된다.
    #   상호 변경·지점 표기 차이가 원인이라 이름을 덮지 않고 **별칭만 추가**한다.
    new, hit_code, hit_name, dup_in_excel = [], 0, 0, 0
    aliases = []
    seen_key = set()
    for r in rows:
        if r['code2'] and r['code2'] in by_code:
            hit_code += 1
            m = by_code[r['code2']]
            k = nrm(r['name'])
            if k != nrm(m['client_name']) and k not in alias_of.get(m['id'], set()):
                aliases.append({'id': m['id'], 'mes_name': m['client_name'], 'ec_name': r['name'],
                                'kws': str(m.get('search_keywords') or '')})
                alias_of.setdefault(m['id'], set()).add(k)
            continue
        if nrm(r['name']) in by_name:
            hit_name += 1
            continue
        key = nrm(r['name'])
        if key in seen_key:          # 엑셀 자체 중복(같은 이름 다른 코드)
            dup_in_excel += 1
            continue
        seen_key.add(key)
        new.append(r)

    print(f'  기존 매칭: 사업자번호 {hit_code} · 이름·별칭 {hit_name} · 엑셀내 중복 {dup_in_excel}')
    print(f'  ▶ 신규 {len(new)}곳 · 별칭 추가 필요 {len(aliases)}곳')

    if aliases:
        print('\n별칭 후보 (사업자번호 동일 · 상호 불일치) — 상위 20')
        for x in aliases[:20]:
            print(f'   #{x["id"]:<6} MES「{x["mes_name"][:22]}」 ← 이카운트「{x["ec_name"][:22]}」')

    # ── 코드 분류(주민번호형은 저장 금지) ────────────────────────────────
    LABEL = {'BIZNO': '사업자번호', 'ECOUNT': '이카운트 내부코드', 'RRN': '주민번호형(대체코드)',
             'OTHER': '코드 무효(대체코드)', 'NONE': '코드없음(대체코드)'}
    kinds = Counter(r['kind'] for r in new)
    for k, v in kinds.most_common():
        print(f'     {LABEL[k]}: {v}')

    print(f'\n신규 {len(new)}곳 전체')
    for r in new:
        shown = r['code2'] if r['kind'] in ('BIZNO', 'ECOUNT') else f'({LABEL[r["kind"]]})'
        print(f'   {shown:<20} {r["name"][:26]:<28} {r["rep"][:10]}')

    if not a.out:
        print('\n(대조만 수행 — SQL 을 만들려면 --out 지정)')
        return

    # ── INSERT SQL ────────────────────────────────────────────────────────
    # client_code 는 UNIQUE NOT NULL. 사업자번호가 있으면 그것을, 없거나 주민번호형이면 대체코드.
    used = set(by_code.keys())
    lines = [
        '-- 이카운트 거래처등록 → MES clients 신규 INSERT',
        f'-- 원천: {os.path.basename(a.src)} · 사용구분 YES 만 · 신규 {len(new)}곳',
        '-- 주민번호형 거래처코드는 저장하지 않는다(대체코드 C-xxxxx, 사업자번호 칸 NULL).',
        '-- ⚠️ BEGIN TRANSACTION/COMMIT 을 넣지 않는다 — D1 이 거부한다.',
        '--    `wrangler d1 execute --file` 자체가 원자적이라(실패 시 원상복구) 필요도 없다.',
    ]
    seq = 0
    for r in new:
        code, bizno = r['code2'], r['bizno']
        if not code or code in used:      # 주민번호형·무효코드·충돌 → 대체코드
            seq += 1
            while f'C-{seq:05d}' in used:
                seq += 1
            code = f'C-{seq:05d}'
        used.add(code)
        lines.append(
            'INSERT INTO clients (client_code, client_name, representative, phone, mobile, '
            'search_keywords, transfer_info, business_registration_number, client_type, is_active, notes) '
            f'VALUES ({q(code)}, {q(r["name"])}, {q(r["rep"])}, {q(r["phone"])}, {q(r["mobile"])}, '
            f'{q(r["keywords"])}, {q(r["transfer"])}, {q(bizno)}, \'SALES\', 1, '
            "'이카운트 거래처목록 26.08.13 동기화');"
        )
    # 별칭 — 이름을 **덮지 않고** search_keywords 에 덧붙인다.
    #   importer(cidx)가 별칭을 읽으므로 이것만으로 매칭이 붙는다. 상호 변경 판단은 사람 몫으로 남긴다.
    if aliases:
        lines.append('')
        lines.append(f'-- 별칭 {len(aliases)}건 (사업자번호 동일 · 상호 불일치) — 기존 이름 유지, 별칭만 추가')
        for x in aliases:
            cur = [k.strip() for k in x['kws'].replace(',', '|').split('|') if k.strip()]
            cur.append(x['ec_name'])
            merged = '|'.join(dict.fromkeys(cur))     # 순서 유지 중복 제거
            lines.append(
                f'UPDATE clients SET search_keywords = {q(merged)}, updated_at = CURRENT_TIMESTAMP '
                f'WHERE id = {int(x["id"])};   -- {x["mes_name"]} ← {x["ec_name"]}'
            )

    with open(a.out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')
    print(f'\nSQL 생성: {a.out} — INSERT {len(new)} · 별칭 UPDATE {len(aliases)}')


if __name__ == '__main__':
    main()
